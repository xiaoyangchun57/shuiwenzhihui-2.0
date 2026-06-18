"""
水利运维智慧运营平台 - 投标演示后端
Flask RESTful API + SQLite + APScheduler

==================== 完整API端点文档 ====================

【系统与概览】
  GET  /api/health
       健康检查。返回 {'status':'ok','time':'...'}

  GET  /api/dashboard/summary
       仪表盘汇总。返回总览数据、最新告警TOP5、待处理工单TOP5、今日巡检统计

  GET  /api/data/overview
       数据概览。返回站点数/在线数、设备数/在线数、活跃告警数、进行中工单数

【站点管理】
  GET  /api/sites
       所有站点列表。返回站点基本信息及设备数量

  GET  /api/sites/<site_id>
       单个站点详情。返回站点信息、设备列表、活跃告警数、进行中工单数

【实时数据】
  GET  /api/data/realtime
       各站点最新传感器数据。返回每个站点的最新一条数据（metric/value/unit/time）

  GET  /api/data/site/<site_id>?limit=50
       指定站点最近N条数据（默认50条）。返回metric/value/unit/recorded_at数组

【告警管理】
  GET  /api/alerts?status=pending&limit=50
       告警列表。支持按status筛选（pending/acknowledged/resolved），按等级priority排序

  POST /api/alerts/<alert_id>/acknowledge
       确认告警。将告警状态改为acknowledged

  GET  /api/alerts/statistics
       告警统计。返回total、by_level(red/orange/yellow/blue)、by_status(pending/acknowledged/resolved)

【工单管理】
  GET  /api/workorders?status=pending&limit=50
       工单列表。支持按status筛选，返回工单及关联站点名

  POST /api/workorders
       创建工单。请求体JSON: {site_id,source,event_type,level,title,description,images,assignee}
       自动生成工单号和SLA截止时间，返回{'success':True,'order_no':'WO-...'}

  PUT  /api/workorders/<order_no>/status
       更新工单状态。请求体JSON: {status,remark?,satisfaction?}
       状态流转：pending->accepted->generated->dispatched->in_progress->reviewing->acceptance->closed

  GET  /api/workorders/statistics
       工单统计。返回total、by_status各状态计数、today_new、today_closed

【巡检管理】
  GET  /api/inspections
       巡检计划列表。返回计划及完成进度(total_items/completed_items)

  POST /api/inspections
       创建巡检计划。请求体JSON: {plan_name,site_id,type,start_date,end_date,check_items?}
       自动生成检查任务项，返回{'success':True,'plan_id':N}

  GET  /api/inspections/<plan_id>/tasks
       巡检任务列表。返回指定计划下所有检查项

  PUT  /api/inspections/tasks/<task_id>
       提交巡检结果。请求体JSON: {result,photo?,gps_lat?,gps_lng?,check_time?,remark?}

  GET  /api/inspections/statistics
       巡检统计。返回计划数/完成数、任务数/完成数、异常数

【热线管理】
  GET  /api/hotline/events?limit=50
       热线事件列表。返回热线来电记录

  POST /api/hotline/events
       登记热线事件。请求体JSON: {caller_name,caller_phone,event_type,description,location,operator}

  POST /api/hotline/events/<event_id>/convert
       热线转工单。请求体JSON: {level,assignee}
       自动生成工单并更新热线事件状态，返回{'success':True,'order_no':'WO-...'}

【天气数据】 -- NEW
  GET  /api/weather
       天气数据。返回当前天气（温度/湿度/风速/风向/降水量/气压/天气类型）、
       未来24小时逐小时预报数组、天气预警列表（暴雨/大风/高温）

【水质监测】 -- NEW
  GET  /api/water-quality?site_id=<可选>
       水质监测数据。返回供水站/水库的水质指标（浊度/pH/余氯/氨氮/COD），
       每个指标含当前值、记录时间和7日均值对比。支持按site_id筛选

【设备监控】 -- NEW
  GET  /api/devices/status
       设备状态汇总。返回设备总数/在线数/离线数、各类型设备统计、离线设备详情列表

【数据质量】 -- NEW
  GET  /api/data-quality
       数据质量报告。返回今日数据到达率/完整率/及时率、异常站点列表、最近24小时质量趋势
"""
import os
import json
import sqlite3
import random
import time
import threading
import hashlib
import secrets
from datetime import datetime, timedelta
from contextlib import contextmanager

from flask import Flask, jsonify, request, g, send_from_directory, send_file
from flask_cors import CORS
from apscheduler.schedulers.background import BackgroundScheduler
import os, uuid, urllib.request, urllib.error, json as _json

app = Flask(__name__, static_folder=None)  # 禁用默认static，手动控制
CORS(app)

DB_PATH = os.path.join(os.path.dirname(__file__), 'data', 'water.db')
scheduler = BackgroundScheduler()
scheduler.start()

# ===================== 状态跟踪（每个站点的当前监测值趋势） =====================

_site_state = {}
def get_site_trend(site_id, metric, base, var, min_v=None, max_v=None):
    """生成有趋势的传感器数据：当前值 = 前值 + 随机趋势，避免纯随机跳变"""
    key = (site_id, metric)
    if key not in _site_state:
        _site_state[key] = base
    was = _site_state[key]
    drift = random.uniform(-var, var)
    val = round(was + drift, 2)
    if min_v is not None:
        val = max(min_v, val)
    if max_v is not None:
        val = min(max_v, val)
    _site_state[key] = val
    return val

# ===================== Database =====================

@contextmanager
def get_db():
    db = sqlite3.connect(DB_PATH, timeout=3, check_same_thread=False)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA busy_timeout=3000")
    db.execute("PRAGMA synchronous=NORMAL")
    db.execute("PRAGMA cache_size=-8000")
    try:
        yield db
    finally:
        db.close()

def init_db():
    with get_db() as db:
        db.executescript('''
            CREATE TABLE IF NOT EXISTS sites (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                type TEXT NOT NULL,
                lat REAL, lng REAL,
                district TEXT DEFAULT '',
                river TEXT DEFAULT '',
                status TEXT DEFAULT 'online',
                manager TEXT, phone TEXT,
                last_heartbeat TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS sensor_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                metric TEXT NOT NULL,
                value REAL NOT NULL,
                unit TEXT,
                threshold_low REAL,
                threshold_high REAL,
                threshold_critical REAL,
                recorded_at TEXT NOT NULL,
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            CREATE TABLE IF NOT EXISTS alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                metric TEXT NOT NULL,
                value REAL NOT NULL,
                level TEXT NOT NULL,  -- blue/yellow/orange/red
                message TEXT NOT NULL,
                status TEXT DEFAULT 'pending',  -- pending/acknowledged/resolved
                created_at TEXT DEFAULT (datetime('now','localtime')),
                resolved_at TEXT,
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            CREATE TABLE IF NOT EXISTS work_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no TEXT UNIQUE NOT NULL,
                site_id INTEGER,
                source TEXT NOT NULL,  -- auto/patrol/hotline/superior
                event_type TEXT NOT NULL,
                level TEXT NOT NULL,  -- normal/urgent/critical
                title TEXT NOT NULL,
                description TEXT,
                images TEXT,
                assignee TEXT,
                status TEXT DEFAULT 'pending',  -- pending/accepted/generated/dispatched/in_progress/reviewing/acceptance/closed
                sla_deadline TEXT,
                resolved_at TEXT,
                remark TEXT,
                satisfaction INTEGER,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            CREATE TABLE IF NOT EXISTS inspection_plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_name TEXT NOT NULL,
                site_id INTEGER NOT NULL,
                type TEXT NOT NULL,  -- daily/weekly/monthly/special
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                status TEXT DEFAULT 'pending',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );
            -- Migrate: add period column if not exists
            -- This ALTER TABLE is executed separately below, not in executescript

            CREATE TABLE IF NOT EXISTS inspection_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plan_id INTEGER NOT NULL,
                site_id INTEGER NOT NULL,
                inspector TEXT,
                check_item TEXT NOT NULL,
                result TEXT,  -- normal/abnormal/na
                photo TEXT,
                gps_lat REAL, gps_lng REAL,
                check_time TEXT,
                remark TEXT,
                FOREIGN KEY (plan_id) REFERENCES inspection_plans(id),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            CREATE TABLE IF NOT EXISTS hotline_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                caller_name TEXT,
                caller_phone TEXT,
                event_type TEXT NOT NULL,
                description TEXT NOT NULL,
                location TEXT,
                status TEXT DEFAULT 'registered',  -- registered/dispatched/closed
                related_order_no TEXT,
                operator TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            CREATE TABLE IF NOT EXISTS device_shadows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                device_code TEXT UNIQUE NOT NULL,
                device_name TEXT NOT NULL,
                device_type TEXT,
                status TEXT DEFAULT 'online',
                battery REAL,
                voltage REAL DEFAULT 0,
                last_data_time TEXT,
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 天气数据表 (新增)
            CREATE TABLE IF NOT EXISTS weather_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                temperature REAL,
                humidity REAL,
                wind_speed REAL,
                wind_direction TEXT,
                precipitation REAL,
                pressure REAL,
                weather_type TEXT,
                warning_info TEXT,
                recorded_at TEXT DEFAULT (datetime('now','localtime'))
            );

            -- 运维计划表
            CREATE TABLE IF NOT EXISTS maintenance_plans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                plan_name TEXT NOT NULL,
                category TEXT NOT NULL,
                frequency TEXT NOT NULL,
                due_date TEXT,
                status TEXT DEFAULT 'pending',
                assignee TEXT,
                completed_at TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 运维计划模板表
            CREATE TABLE IF NOT EXISTS maintenance_templates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT NOT NULL,
                sub_category TEXT NOT NULL,
                title TEXT NOT NULL,
                frequency TEXT NOT NULL,
                description TEXT,
                standard TEXT,
                check_items TEXT,
                photo_required INTEGER DEFAULT 0,
                estimated_hours REAL DEFAULT 1,
                sort_order INTEGER DEFAULT 0
            );

            -- 数据到报表
            CREATE TABLE IF NOT EXISTS data_arrival (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                metric TEXT NOT NULL,
                expected_count INTEGER DEFAULT 0,
                actual_count INTEGER DEFAULT 0,
                arrival_rate REAL DEFAULT 0,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 水位校验表
            CREATE TABLE IF NOT EXISTS water_level_checks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                manual_level REAL,
                telemetry_level REAL,
                diff REAL,
                status TEXT DEFAULT 'normal',
                adjust_action TEXT,
                operator TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 时间线事件表
            CREATE TABLE IF NOT EXISTS timeline_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_type TEXT NOT NULL,
                source_id INTEGER NOT NULL,
                event_type TEXT NOT NULL,
                operator TEXT,
                remark TEXT,
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            -- 用户表（登录系统）
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'operator',
                real_name TEXT NOT NULL,
                phone TEXT DEFAULT '',
                status TEXT DEFAULT 'active',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            );

            -- 用户-站点分配（多对多）
            CREATE TABLE IF NOT EXISTS user_sites (
                user_id INTEGER NOT NULL,
                site_id INTEGER NOT NULL,
                PRIMARY KEY (user_id, site_id),
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );
            -- 巡检方案母表
            CREATE TABLE IF NOT EXISTS inspection_schemes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                site_id INTEGER NOT NULL,
                period TEXT NOT NULL,
                name TEXT NOT NULL,
                status TEXT DEFAULT 'active',
                updated_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id),
                UNIQUE(site_id, period)
            );

            -- 方案检查项明细
            CREATE TABLE IF NOT EXISTS inspection_scheme_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                scheme_id INTEGER NOT NULL,
                category TEXT DEFAULT '',
                check_item TEXT NOT NULL,
                sort_order INTEGER DEFAULT 0,
                is_required INTEGER DEFAULT 1,
                FOREIGN KEY (scheme_id) REFERENCES inspection_schemes(id) ON DELETE CASCADE
            );

            -- 备件库存表
            CREATE TABLE IF NOT EXISTS spare_parts_inventory (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                part_code TEXT UNIQUE NOT NULL,
                part_name TEXT NOT NULL,
                category TEXT DEFAULT '其他',
                unit TEXT DEFAULT '个',
                quantity INTEGER DEFAULT 0,
                min_quantity INTEGER DEFAULT 5,
                site_id INTEGER,
                remark TEXT DEFAULT '',
                updated_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 备件申请表
            CREATE TABLE IF NOT EXISTS spare_part_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_no TEXT UNIQUE NOT NULL,
                site_id INTEGER NOT NULL,
                applicant TEXT NOT NULL,
                part_name TEXT NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 1,
                reason TEXT DEFAULT '',
                status TEXT DEFAULT 'pending',
                approver TEXT DEFAULT '',
                approval_comment TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                updated_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (site_id) REFERENCES sites(id)
            );

            -- 库存变更流水表
            CREATE TABLE IF NOT EXISTS inventory_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                part_id INTEGER NOT NULL,
                type TEXT NOT NULL CHECK(type IN ('in','out')),
                quantity INTEGER NOT NULL,
                ref_type TEXT DEFAULT '',
                ref_id INTEGER,
                operator TEXT DEFAULT '',
                remark TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime')),
                FOREIGN KEY (part_id) REFERENCES spare_parts_inventory(id)
            );
        ''')
        # 兼容已有数据库：尝试添加列，忽略已存在的错误
        for col_sql in [
            "ALTER TABLE alerts ADD COLUMN urge_count INTEGER DEFAULT 0",
            "ALTER TABLE alerts ADD COLUMN last_urged_at TEXT",
            "ALTER TABLE alerts ADD COLUMN related_order_no TEXT",
            "ALTER TABLE alerts ADD COLUMN response_deadline TEXT",
            "ALTER TABLE maintenance_plans ADD COLUMN urge_count INTEGER DEFAULT 0",
            "ALTER TABLE maintenance_plans ADD COLUMN last_urged_at TEXT",
            "ALTER TABLE maintenance_plans ADD COLUMN review_status TEXT DEFAULT NULL",
            "ALTER TABLE maintenance_plans ADD COLUMN review_comment TEXT DEFAULT NULL",
            "ALTER TABLE maintenance_plans ADD COLUMN template_id INTEGER",
            "ALTER TABLE maintenance_plans ADD COLUMN sub_category TEXT",
            "ALTER TABLE maintenance_plans ADD COLUMN check_results TEXT",
            "ALTER TABLE maintenance_plans ADD COLUMN remark TEXT",
            "ALTER TABLE inspection_plans ADD COLUMN period TEXT DEFAULT 'once'",
            "ALTER TABLE inspection_plans ADD COLUMN description TEXT DEFAULT ''",
            "ALTER TABLE inspection_plans ADD COLUMN scheme_id INTEGER",
            "ALTER TABLE device_shadows ADD COLUMN voltage REAL DEFAULT 0",
        ]:
            try:
                db.execute(col_sql)
            except:
                pass
        db.commit()

# ===================== Seed Data =====================

_NC_CENTER = (28.68, 115.86)  # 南昌市中心

def _gen_nanchang_sites():
    """生成南昌市300+水文监测站点"""
    sites = []
    # 赣江干流关键节点（南→北）
    ganjiang = [
        (28.20, 115.95), (28.24, 115.93), (28.28, 115.90), (28.32, 115.88),
        (28.36, 115.86), (28.40, 115.85), (28.44, 115.84), (28.48, 115.85),
        (28.52, 115.87), (28.56, 115.90), (28.60, 115.93), (28.64, 115.96),
        (28.68, 115.99), (28.72, 116.02), (28.76, 116.04), (28.80, 116.06),
        (28.84, 116.08), (28.88, 116.10), (28.92, 116.12),
    ]
    # 抚河关键节点（东南→西北）
    fuhe = [
        (28.30, 116.20), (28.34, 116.16), (28.38, 116.12), (28.42, 116.08),
        (28.46, 116.04), (28.50, 116.00), (28.54, 115.96),
    ]
    # 鄱阳湖沿岸（东北）
    poyang = [
        (28.90, 116.08), (28.95, 116.12), (29.00, 116.16), (29.05, 116.20),
        (29.10, 116.18), (29.15, 116.14),
    ]
    # 区县域坐标范围
    districts = {
        '西湖': (28.65, 115.87), '东湖': (28.69, 115.89), '青山湖': (28.70, 115.95),
        '青云谱': (28.63, 115.92), '新建': (28.70, 115.82), '南昌县': (28.55, 115.95),
        '进贤': (28.38, 116.24), '安义': (28.85, 115.55), '湾里': (28.72, 115.73),
    }
    sid = 0
    # 生成雨量站（120个）：沿河高密度+区域均匀
    for i, (blat, blng) in enumerate(ganjiang):
        for j in range(4):  # 每段4个
            lat = blat + random.uniform(-0.04, 0.04)
            lng = blng + random.uniform(-0.04, 0.04)
            sid += 1; dname = random.choice(list(districts.keys()))
            sites.append((f'YL-{dname[:2].upper()}-{sid:03d}', f'{dname}雨量站{sid}', 'rainfall', round(lat,4), round(lng,4), dname, '赣江'))
    # 沿抚河补充
    for i, (blat, blng) in enumerate(fuhe):
        for j in range(3):
            lat = blat + random.uniform(-0.03, 0.03)
            lng = blng + random.uniform(-0.03, 0.03)
            sid += 1; dname = random.choice(['南昌县','进贤'])
            sites.append((f'YL-{dname[:2].upper()}-{sid:03d}', f'{dname}雨量站{sid}', 'rainfall', round(lat,4), round(lng,4), dname, '抚河'))
    # 市区低密度补充到120个
    while len([s for s in sites if s[2]=='rainfall']) < 120:
        lat = random.uniform(28.45, 28.95); lng = random.uniform(115.50, 116.40)
        # 避免离已有站点太近
        too_close = any(abs(s[3]-lat)+abs(s[4]-lng)<0.06 for s in sites)
        if not too_close:
            sid += 1
            dname = min(districts.keys(), key=lambda d: abs(lat-districts[d][0])+abs(lng-districts[d][1]))
            sites.append((f'YL-{dname[:2].upper()}-{sid:03d}', f'{dname}雨量站{sid}', 'rainfall', round(lat,4), round(lng,4), dname, ''))

    # 生成水位站（90个）：赣江沿岸高密度
    for i, (blat, blng) in enumerate(ganjiang):
        for j in range(3):
            lat = blat + random.uniform(-0.015, 0.015)
            lng = blng + random.uniform(-0.015, 0.015)
            sid += 1; dname = random.choice(list(districts.keys()))
            sites.append((f'SW-{dname[:2].upper()}-{sid:03d}', f'{dname}水位站{sid}', 'water_level', round(lat,4), round(lng,4), dname, '赣江'))
    # 抚河补充
    for i, (blat, blng) in enumerate(fuhe):
        for j in range(2):
            lat = blat + random.uniform(-0.015, 0.015)
            lng = blng + random.uniform(-0.015, 0.015)
            sid += 1; dname = random.choice(['南昌县','进贤'])
            sites.append((f'SW-{dname[:2].upper()}-{sid:03d}', f'{dname}水位站{sid}', 'water_level', round(lat,4), round(lng,4), dname, '抚河'))
    # 鄱阳湖补充
    for i, (blat, blng) in enumerate(poyang):
        sid += 1
        sites.append((f'SW-PY-{sid:03d}', f'鄱阳水位站{sid}', 'water_level', round(blat+random.uniform(-0.02,0.02),4), round(blng+random.uniform(-0.02,0.02),4), '进贤', '鄱阳湖'))

    # 生成水文站（45个）：关键断面
    key_points = ganjiang[::2] + fuhe[::2] + poyang[::2]
    for i, (blat, blng) in enumerate(key_points):
        for j in range(2 if i < 8 else 1):
            lat = blat + random.uniform(-0.01, 0.01); lng = blng + random.uniform(-0.01, 0.01)
            sid += 1; dname = random.choice(list(districts.keys()))
            river = '赣江' if i < len(ganjiang)//2 else ('抚河' if i < len(ganjiang)//2+len(fuhe)//2 else '鄱阳湖')
            sites.append((f'HW-{dname[:2].upper()}-{sid:03d}', f'{dname}水文站{sid}', 'hydrology', round(lat,4), round(lng,4), dname, river))

    # 生成墒情站（30个）：农田/灌区
    farm_areas = [(28.45,115.85),(28.50,115.90),(28.55,115.92),(28.60,115.88),(28.65,115.80),
                  (28.70,115.78),(28.75,115.85),(28.40,116.10),(28.45,116.05),(28.35,116.15)]
    for i in range(30):
        if i < len(farm_areas):
            lat, lng = farm_areas[i]
        else:
            lat = random.uniform(28.35,28.80); lng = random.uniform(115.60,116.20)
        lat += random.uniform(-0.02,0.02); lng += random.uniform(-0.02,0.02)
        sid += 1; dname = min(districts.keys(), key=lambda d: abs(lat-districts[d][0])+abs(lng-districts[d][1]))
        sites.append((f'SQ-{dname[:2].upper()}-{sid:03d}', f'{dname}墒情站{sid}', 'soil_moisture', round(lat,4), round(lng,4), dname, ''))

    # 生成蒸发站（15个）：空旷地带
    open_areas = [(28.72,115.73),(28.80,115.60),(28.60,115.70),(28.50,115.75),
                  (28.90,115.90),(28.75,116.00),(28.40,115.88),(28.55,115.65),
                  (28.70,115.55),(28.85,115.70),(28.45,115.78),(28.62,115.82),
                  (28.92,115.95),(28.52,115.68),(28.78,115.92)]
    for i, (lat, lng) in enumerate(open_areas):
        sid += 1; dname = min(districts.keys(), key=lambda d: abs(lat-districts[d][0])+abs(lng-districts[d][1]))
        sites.append((f'ZF-{dname[:2].upper()}-{sid:03d}', f'{dname}蒸发站{sid}', 'evaporation', round(lat+random.uniform(-0.01,0.01),4), round(lng+random.uniform(-0.01,0.01),4), dname, ''))

    return sites

def seed_data():
    """种子数据：真实站点 + 设备 + 工单 + 热线事件（仅首次运行）"""
    with get_db() as db:
        count = db.execute("SELECT COUNT(*) FROM sites").fetchone()[0]
        if count > 0:
            print("[Seed] 站点数据已存在，跳过站点/设备/工单种子数据")
            return

        # === 235个真实站点导入 ===
        import json as _json
        json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'site_data.json')
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                all_sites = _json.load(f)
        else:
            all_sites = _gen_nanchang_sites()
        for s in all_sites:
            lat = s['lat'] or random.uniform(28.4, 29.2)
            lng = s['lng'] or random.uniform(115.5, 116.5)
            db.execute(
                "INSERT INTO sites (code,name,type,lat,lng,district) VALUES (?,?,?,?,?,?)",
                (s['code'], s['name'], s['type'], lat, lng, s.get('address',''))
            )
        print(f"[Seed] 生成 {len(all_sites)} 个站点")

        # === 分配负责人（同一行政区划分给同一运维人员，与人员管理一致） ===
        real_users = db.execute("SELECT username, real_name FROM users WHERE role='operator' ORDER BY id").fetchall()
        real_names = [u['real_name'] for u in real_users]
        if not real_names: real_names = ['张建国','黎明','王刚','赵洪']
        all_rows = db.execute("SELECT id, district FROM sites ORDER BY district, id").fetchall()
        mgr_map = {}; mgr_idx = 0
        for row in all_rows:
            dist = row['district'] or ''
            if dist not in mgr_map:
                mgr_map[dist] = real_names[mgr_idx % len(real_names)]
                mgr_idx += 1
            db.execute("UPDATE sites SET manager=?, phone=? WHERE id=?",
                       (mgr_map[dist], f'1{random.randint(30,39)}0000{random.randint(1000,9999)}', row['id']))

        # === 设备生成（每站按类型配设备） ===
        type_devices = {
            'rainfall': [('翻斗式雨量计','rainfall_gauge'),('电子雨量计','electronic_rainfall')],
            'water_level': [('雷达水位计','radar_water_level'),('压力式水位计','pressure_water_level'),('流速计','flow_meter')],
            'hydrology': [('水文综合采集仪','hydro_collector'),('流速仪','current_meter'),('雨量计','rainfall_meter'),('水位计','water_level_meter')],
            'soil_moisture': [('土壤水分传感器','soil_moisture_sensor'),('土壤温度计','soil_temperature')],
            'evaporation': [('蒸发皿','evaporation_pan'),('气象百叶箱','weather_screen'),('风速仪','anemometer')],
            'groundwater': [('地下水位计','groundwater_level'),('水质在线监测仪','water_quality_monitor')],
            'station_yard': [('视频监控','video_surveillance'),('安防报警','security_alarm'),('环境传感器','env_sensor')],
        }
        all_sites_db = db.execute("SELECT id, code, type FROM sites ORDER BY id").fetchall()
        for site in all_sites_db:
            devs = type_devices.get(site['type'], [('通用传感器','generic')])
            for i, (dname, dtype) in enumerate(devs):
                db.execute(
                    "INSERT INTO device_shadows (site_id,device_code,device_name,device_type,status,battery,voltage) VALUES (?,?,?,?,?,?,?)",
                    (site['id'], f"{site['code']}-{i+1:02d}{dtype[:4].upper()}", dname, dtype,
                     'online', round(random.uniform(60,100), 0),
                     round(random.uniform(11.5, 14.2), 1))
                )

        # 工单种子数据（取前几个站ID）
        sample_ids = [r['id'] for r in db.execute("SELECT id FROM sites ORDER BY id LIMIT 5").fetchall()]
        orders = [
            (f'WO-20260618-{i+1:03d}', sample_ids[i] if i < len(sample_ids) else sample_ids[0],
             'auto','设备故障','normal','水位计数据中断','设备持续30分钟无数据上报','', '张建国','dispatched','2026-06-18 16:00','2026-06-18 08:30') for i in range(5)
        ]
        for o in orders:
            db.execute(
                "INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,images,assignee,status,sla_deadline,created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                o
            )

        # 热线事件
        hotlines = [
            ('张先生','13900001100','水质问题','家里的自来水发黄，已经持续两天了','城南街道阳光小区','registered', '', '李敏','2026-06-10 14:05'),
            ('李女士','13900002200','设施损坏','河道护栏被撞坏，存在安全隐患','滨江路新华桥东侧','dispatched','WO-20260611-004','李敏','2026-06-10 16:30'),
            ('陈先生','13900003300','违规举报','有人在河道内非法采砂','滨江堤防B段下游','registered', '', '王芳','2026-06-11 08:15'),
            ('匿名','', '水位异常','东湖水位这两天涨得很快，担心漫堤','东湖公园湖区','registered', '', '王芳','2026-06-11 09:30'),
        ]
        for h in hotlines:
            db.execute(
                "INSERT INTO hotline_events (caller_name,caller_phone,event_type,description,location,status,related_order_no,operator,created_at) VALUES (?,?,?,?,?,?,?,?,?)",
                h
            )

        # 备件库存种子数据
        inv_cnt = db.execute("SELECT COUNT(*) FROM spare_parts_inventory").fetchone()[0]
        if inv_cnt == 0:
            spare_parts = [
                ('BJ-001', '水位计传感器', '传感器', '个', 30, 5),
                ('BJ-002', '雨量筒翻斗', '传感器', '个', 15, 5),
                ('BJ-003', '太阳能板(20W)', '电源', '块', 8, 3),
                ('BJ-004', '蓄电池(12V)', '电源', '个', 12, 5),
                ('BJ-005', '数据采集终端RTU', '通信', '台', 5, 2),
                ('BJ-006', 'GPRS通信模块', '通信', '个', 10, 3),
                ('BJ-007', '不锈钢水位计支架', '结构', '套', 6, 3),
                ('BJ-008', '防雷模块', '电源', '个', 20, 5),
                ('BJ-009', '信号线缆(10m)', '线缆', '根', 25, 10),
                ('BJ-010', '水位计密封圈', '其他', '个', 50, 10),
                ('BJ-011', '温湿度传感器', '传感器', '个', 8, 3),
                ('BJ-012', '风速风向仪', '传感器', '台', 3, 2),
            ]
            for pc, pn, cat, unit, qty, minq in spare_parts:
                db.execute(
                    "INSERT INTO spare_parts_inventory (part_code,part_name,category,unit,quantity,min_quantity) VALUES (?,?,?,?,?,?)",
                    (pc, pn, cat, unit, qty, minq)
                )

        # 备件申请种子数据（演示用）
        req_cnt = db.execute("SELECT COUNT(*) FROM spare_part_requests").fetchone()[0]
        if req_cnt == 0:
            from datetime import datetime, timedelta
            now = datetime.now()
            sample_reqs = [
                (1, '系统管理员', '水位计传感器', 2, '水位计数据异常，需更换', 'approved', '系统管理员', '同意更换', (now - timedelta(hours=2)).strftime('%Y-%m-%d %H:%M:%S')),
                (2, '系统管理员', '太阳能板(20W)', 1, '太阳能板破损', 'approved', '系统管理员', '已核实，批准', (now - timedelta(hours=5)).strftime('%Y-%m-%d %H:%M:%S')),
                (3, '运维人员', 'GPRS通信模块', 2, '通信模块频繁断连', 'pending', '', '', (now - timedelta(minutes=30)).strftime('%Y-%m-%d %H:%M:%S')),
                (4, '运维人员', '防雷模块', 3, '汛期前补充', 'pending', '', '', (now - timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')),
                (5, '系统管理员', '数据采集终端RTU', 1, 'RTU老化需更换', 'rejected', '系统管理员', '库存不足，暂缓采购', (now - timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')),
            ]
            for idx, (sid, applicant, pname, qty, reason, status, approver, comment, ctime) in enumerate(sample_reqs):
                rno = f"BJ-{now.strftime('%Y%m%d')}-{idx+1:03d}"
                db.execute(
                    "INSERT INTO spare_part_requests (request_no,site_id,applicant,part_name,quantity,reason,status,approver,approval_comment,created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    (rno, sid, applicant, pname, qty, reason, status, approver, comment, ctime)
                )
                # 已批准的申请记录扣减库存流水
                if status == 'approved':
                    inv = db.execute("SELECT id, quantity FROM spare_parts_inventory WHERE part_name LIKE ? LIMIT 1", (f"%{pname}%",)).fetchone()
                    if inv:
                        new_qty = max(0, inv['quantity'] - qty)
                        db.execute("UPDATE spare_parts_inventory SET quantity=? WHERE id=?", (new_qty, inv['id']))
                        db.execute("INSERT INTO inventory_logs (part_id,type,quantity,ref_type,ref_id,operator,remark) VALUES (?,'out',?,'request',?,?,?)",
                            (inv['id'], qty, 0, '系统管理员', f"种子数据：{rno}"))

        db.commit()
        print("[Seed] Database seeded with initial data.")

def seed_inspections():
    """巡检种子数据（独立判断，可重复运行）"""
    with get_db() as db:
        cnt = db.execute("SELECT COUNT(*) FROM inspection_plans").fetchone()[0]
        if cnt == 0:
                all_sites = db.execute("SELECT id, name, type FROM sites ORDER BY id").fetchall()
                if len(all_sites) < 5: return
                s = all_sites
                insp_plans = [
                    (f'{s[0]["name"]}日常巡检', s[0]['id'],'daily','2026-06-10','2026-06-10','completed'),
                    (f'{s[0]["name"]}周巡检', s[0]['id'],'weekly','2026-06-08','2026-06-14','in_progress'),
                    (f'{s[1]["name"]}日常巡检', s[1]['id'],'daily','2026-06-11','2026-06-11','pending'),
                    (f'{s[2]["name"]}日常巡检', s[2]['id'],'daily','2026-06-10','2026-06-10','completed'),
                    (f'{s[3]["name"]}周巡检', s[3]['id'],'weekly','2026-06-09','2026-06-15','in_progress'),
                    (f'{s[4]["name"]}日常巡检', s[4]['id'],'daily','2026-06-11','2026-06-11','pending'),
                    (f'{s[5]["name"] if len(s) > 5 else s[0]["name"]}日常巡检', s[5]['id'] if len(s) > 5 else s[0]['id'],'daily','2026-06-10','2026-06-10','completed'),
                ]
                insp_items_map = {
                    'reservoir': ['坝体外观检查','溢洪道检查','放水设施检查','监测设备检查','防汛物资检查','管理设施检查'],
                    'sluice': ['闸门启闭检查','电气设备检查','上下游检查','监测设备检查','管理设施检查'],
                    'dike': ['堤身外观检查','堤脚检查','排水设施检查','监测设备检查'],
                    'pump': ['机组运行检查','电气设备检查','管道阀门检查','自动化系统检查'],
                    'water_supply': ['取水口检查','净化设备检查','加药系统检查','水质检测','管理设施检查'],
                }
                site_types = {s['id']: s['type'] for s in db.execute("SELECT id,type FROM sites").fetchall()}
                for plan in insp_plans:
                    pname, psid, ptype, pstart, pend, pstatus = plan
                    cur = db.execute(
                        "INSERT INTO inspection_plans (plan_name,site_id,type,start_date,end_date,status) VALUES (?,?,?,?,?,?)",
                        (pname, psid, ptype, pstart, pend, pstatus)
                    )
                    pid = cur.lastrowid
                    items = insp_items_map.get(site_types.get(psid, 'reservoir'), insp_items_map['reservoir'])
                    for item in items:
                        db.execute(
                            "INSERT INTO inspection_tasks (plan_id,site_id,check_item,result,remark,check_time) VALUES (?,?,?,?,?,?)",
                            (pid, psid, item, 'normal' if pstatus == 'completed' else None,
                             '一切正常' if pstatus == 'completed' else None,
                             pstart + ' 09:00' if pstatus == 'completed' else None)
                        )
                    # 对 in_progress 的计划，部分任务已完成
                    if pstatus == 'in_progress':
                        partial = db.execute(
                            "SELECT id FROM inspection_tasks WHERE plan_id=? LIMIT ?", (pid, len(items)//2)
                        ).fetchall()
                        for r in partial:
                            db.execute(
                                "UPDATE inspection_tasks SET result='normal', remark='运行正常', check_time=? WHERE id=?",
                                (pstart + ' 08:30', r['id'])
                            )
        db.commit()
        print("[Seed] Inspection plans seeded.")

def seed_alerts():
    """历史告警种子数据（仅首次）"""
    with get_db() as db:
        acnt = db.execute("SELECT COUNT(*) FROM alerts").fetchone()[0]
        if acnt == 0:
            # 从各类型站点取前几个生成告警
            sid_map = {}
            for t in ['water_level','rainfall','hydrology']:
                r = db.execute("SELECT id FROM sites WHERE type=? ORDER BY id LIMIT 5", (t,)).fetchall()
                if r: sid_map[t] = [row['id'] for row in r]
            def pick(t, idx=0):
                lst = sid_map.get(t, [1])
                return lst[idx % len(lst)]
            alerts_seed = [
                (pick('water_level',0),'water_level',50.8,'yellow','acknowledged','库水位偏高 50.8m > 49.5m','2026-06-11 06:15:00'),
                (pick('hydrology',0),'displacement',11.2,'orange','acknowledged','位移超限！11.2mm > 10.0mm','2026-06-11 07:30:00'),
                (pick('water_level',1),'vibration',8.5,'yellow','acknowledged','振动超限！8.5mm/s > 7.0mm/s','2026-06-11 07:45:00'),
                (pick('rainfall',0),'turbidity',0.78,'red','acknowledged','浊度超标！0.78NTU > 0.5NTU','2026-06-11 08:00:00'),
                (pick('water_level',0),'water_level',51.8,'red','acknowledged','水位超危急线！51.8m > 51.5m','2026-06-11 09:10:00'),
                (pick('rainfall',0),'chlorine',0.55,'orange','pending','余氯偏高！0.55mg/L','2026-06-11 09:30:00'),
                (pick('water_level',1),'vibration',9.3,'orange','pending','振动严重！9.3mm/s','2026-06-11 09:45:00'),
                (pick('water_level',2),'seepage',0.85,'yellow','pending','渗流量偏大 0.85L/s','2026-06-11 10:00:00'),
                (pick('water_level',3),'water_level_upstream',14.8,'yellow','pending','上游水位偏高 14.8m','2026-06-11 10:15:00'),
                (pick('rainfall',1),'ph',8.4,'yellow','pending','pH值异常 8.4','2026-06-11 10:20:00'),
            ]
            for a in alerts_seed:
                sid = a[0]; metric = a[1]; val = a[2]; lv = a[3]; st = a[4]; msg = a[5]; ct = a[6]
                db.execute(
                    "INSERT INTO alerts (site_id,metric,value,level,status,message,created_at,resolved_at) VALUES (?,?,?,?,?,?,?,?)",
                    (sid, metric, val, lv, st, msg, ct, ct if st != 'pending' else None)
                )
            db.commit()
            print("[Seed] Historical alerts seeded.")

def seed_maintenance():
    """运维计划种子数据（仅首次）"""
    with get_db() as db:
        mcnt = db.execute("SELECT COUNT(*) FROM maintenance_plans").fetchone()[0]
        if mcnt == 0:
            sites = db.execute("SELECT id, name FROM sites LIMIT 50").fetchall()
            categories = [
                ('站院环境维护','environment','weekly'),
                ('站房维护','facility','biweekly'),
                ('设施设备维护','facility','monthly'),
                ('观测场管理','observation','biweekly'),
                ('断面环境管理','section','monthly'),
                ('安全检查','safety','monthly'),
                ('发电机保养','generator','monthly'),
            ]
            now = datetime.now()
            import itertools
            plan_id_counter = itertools.count(1)
            for site in sites:
                for cat_name, cat_key, freq in categories:
                    if freq == 'weekly':
                        due = now + timedelta(days=random.randint(0,7))
                    elif freq == 'biweekly':
                        due = now + timedelta(days=random.randint(0,14))
                    elif freq == 'monthly':
                        due = now + timedelta(days=random.randint(0,30))
                    else:
                        due = now + timedelta(days=random.randint(0,90))
                    # 60%概率已处理，40%待处理
                    status = 'completed' if random.random() < 0.6 else 'pending'
                    completed_at = due.strftime('%Y-%m-%d') if status == 'completed' else None
                    db.execute(
                        "INSERT INTO maintenance_plans (site_id,plan_name,category,frequency,due_date,status,assignee,completed_at) VALUES (?,?,?,?,?,?,?,?)",
                        (site['id'], f"{site['name']}{cat_name}", cat_key, freq,
                         due.strftime('%Y-%m-%d'), status, '管理员', completed_at)
                    )
            db.commit()
            print(f"[Seed] {len(sites)*len(categories)} maintenance plans seeded.")

def seed_maintenance_templates():
    """预置标准化运维模板（仅首次）"""
    templates = [
        # === 日常维护类 ===
        ('日常维护','environment','驻测站站院环境维护（每周）','weekly',
         '对水位井、站院、大门口进行全面的打扫，确保干净整洁',
         '地面、窗台、设备等干净整洁，墙面、天花板无污迹、蜘蛛网、昆虫等',
         '[{"id":"c1","label":"水位井区域全面打扫"},{"id":"c2","label":"站院地面清洁"},{"id":"c3","label":"大门口区域打扫"},{"id":"c4","label":"设备表面擦拭"},{"id":"c5","label":"墙面天花板检查（无污迹/蜘蛛网）"}]',
         1, 2, 1),
        ('日常维护','facility','巡测站站房维护（每月2次）','biweekly',
         '对站房进行全面的打扫，确保干净整洁',
         '地面、窗台、设备等干净整洁，墙面、天花板无污迹、蜘蛛网、昆虫等',
         '[{"id":"c1","label":"站房地面清洁"},{"id":"c2","label":"窗台及设备擦拭"},{"id":"c3","label":"墙面天花板检查"},{"id":"c4","label":"门窗完好检查"}]',
         1, 1.5, 2),
        ('日常维护','observation','观测场管理（每月2次）','biweekly',
         '对降蒸观测场草地进行维护，草皮高度符合规范要求',
         '降蒸观测场、站院草皮高度低于20cm，遇重大活动增加维护次数',
         '[{"id":"c1","label":"草地修剪"},{"id":"c2","label":"草高测量（需≤20cm）"},{"id":"c3","label":"杂草清理"},{"id":"c4","label":"场地平整度检查"}]',
         1, 2, 3),
        ('日常维护','section','断面环境管理（每月）','monthly',
         '对测流断面上下游各5米进行清理杂草、杂木，确保断面整洁',
         '断面无积水、无淤泥、无杂草、无杂物',
         '[{"id":"c1","label":"上下游各5米杂草清理"},{"id":"c2","label":"缆道铁塔四周清理"},{"id":"c3","label":"水尺码头淤泥清理"},{"id":"c4","label":"基本水尺底部清理"},{"id":"c5","label":"拍照存档"}]',
         1, 3, 4),
        # === 日常管理类 ===
        ('日常管理','facility','设施设备巡查（每月）','monthly',
         '检查清洗水尺，对设施设备、爬梯、护栏牢固度进行全面检查',
         '填写设施设备巡查表，有异常维修、拍照存档并报中心站网监测科',
         '[{"id":"c1","label":"水尺清洗检查"},{"id":"c2","label":"爬梯牢固度检查"},{"id":"c3","label":"护栏牢固度检查"},{"id":"c4","label":"设施设备外观检查"},{"id":"c5","label":"异常拍照存档"}]',
         1, 2, 5),
        ('日常管理','safety','安全检查（每月）','monthly',
         '对测验设施设备、安全环境、站房、灭火器、安全器材进行一次安全检查',
         '填记好安全检查记录，存在安全隐患需及时告知中心',
         '[{"id":"c1","label":"灭火器压力检查"},{"id":"c2","label":"安全器材完好性"},{"id":"c3","label":"站房结构安全检查"},{"id":"c4","label":"电气线路检查"},{"id":"c5","label":"填写安全检查记录"}]',
         1, 1.5, 6),
        ('日常管理','generator','发电机保养（每月+汛期）','monthly',
         '对发电机进行检查机油、线路等部件，发电运行时间不少于30分钟',
         '每年汛前汛后保养，更换机油及线路，备足燃料及机油',
         '[{"id":"c1","label":"机油液位检查"},{"id":"c2","label":"线路及各部件检查"},{"id":"c3","label":"发电运行≥30分钟"},{"id":"c4","label":"燃料及机油储备检查"},{"id":"c5","label":"记录运行时间"}]',
         1, 1.5, 7),
        # === 设备仪器维护类 ===
        ('设备仪器维护','water_level','水位项目日常巡查（每日）','weekly',
         '观测基本水尺读数并记录，校对遥测水位及时间',
         '人工与遥测水位相差≥0.02m时需校对，以人工观测为准调整',
         '[{"id":"c1","label":"基本水尺读数记录"},{"id":"c2","label":"遥测水位校对"},{"id":"c3","label":"偏差检测（≥0.02m告警）"},{"id":"c4","label":"水尺清洗检查"},{"id":"c5","label":"填记水位巡查表并拍照存档"}]',
         1, 0.5, 8),
        ('设备仪器维护','rainfall','雨量项目日常巡检（每月）','monthly',
         '遥测雨量器现场运行维护巡检，含数据采集终端、供电、雨量筒检查',
         '每季度进行注水试验，误差≤±4%，大暴雨后及时检查',
         '[{"id":"c1","label":"数据采集终端检查"},{"id":"c2","label":"供电设备检查"},{"id":"c3","label":"布线检查"},{"id":"c4","label":"雨量筒外观/水平检查"},{"id":"c5","label":"环境清理"},{"id":"c6","label":"季度注水试验（误差≤±4%）"}]',
         1, 2, 9),
        ('设备仪器维护','evaporation','蒸发项目日常巡检（每月）','monthly',
         '自动蒸发设备遥测终端现场运行维护巡检',
         '每月不少于1次巡测，每半年渗漏检查，每月至少换水一次',
         '[{"id":"c1","label":"自动蒸发设备检查"},{"id":"c2","label":"半年渗漏检查"},{"id":"c3","label":"蒸发器换水"},{"id":"c4","label":"水圈清洁"},{"id":"c5","label":"数据合理性检查"}]',
         1, 1.5, 10),
        ('设备仪器维护','soil_moisture','墒情站日常巡查（季度）','seasonal',
         '对墒情基本站巡查，保持整洁、数据校测',
         '每季度对基本站巡查不少于1次，保持机箱内干净整洁，清理周边杂草',
         '[{"id":"c1","label":"机箱内部清洁"},{"id":"c2","label":"周边杂草清理"},{"id":"c3","label":"无积水检查"},{"id":"c4","label":"数据校测记录"}]',
         0, 1.5, 11),
        ('设备仪器维护','hydrology','缆道日常巡检（测流时）','seasonal',
         '测流时对行主索、循环索、锚碇、导向轮、绞车等进行检查维护',
         '检查锚碇有无位移、钢丝绳夹头是否松动、绞车运转情况',
         '[{"id":"c1","label":"主索/循环索检查"},{"id":"c2","label":"拉线/卡头检查"},{"id":"c3","label":"锚碇检查（位移/生锈）"},{"id":"c4","label":"导向轮/游轮/行车架检查"},{"id":"c5","label":"绞车运转检查"},{"id":"c6","label":"异常拍照留底"}]',
         1, 3, 12),
    ]
    with get_db() as db:
        cnt = db.execute("SELECT COUNT(*) FROM maintenance_templates").fetchone()[0]
        if cnt == 0:
            for t in templates:
                db.execute(
                    "INSERT INTO maintenance_templates (category,sub_category,title,frequency,description,standard,check_items,photo_required,estimated_hours,sort_order) VALUES (?,?,?,?,?,?,?,?,?,?)",
                    t
                )
            db.commit()
            print(f"[Seed] {len(templates)} maintenance templates seeded.")

# ===================== Simulator =====================

# 各河流警戒水位配置
RIVER_THRESHOLDS = {
    '赣江': {'high': 22.0, 'critical': 23.5, 'base': 18.5},
    '抚河': {'high': 32.0, 'critical': 33.5, 'base': 30.0},
    '鄱阳湖': {'high': 18.5, 'critical': 19.8, 'base': 16.5},
    '': {'high': 15.0, 'critical': 16.5, 'base': 13.0},  # 城区默认
}

# 站点类型对应的监测指标
TYPE_METRICS = {
    'rainfall': [
        ('precipitation','mm',20,50),
        ('cumulative_rainfall','mm',100,200),
    ],
    'water_level': [
        ('water_level','m',None,None),  # 阈值在生成时动态计算
        ('flow','m³/s',None,None),
    ],
    'hydrology': [
        ('water_level','m',None,None),
        ('velocity','m/s',None,None),
        ('flow','m³/s',None,None),
        ('precipitation','mm',20,50),
    ],
    'soil_moisture': [
        ('soil_moisture','%',90,None),   # 90%渍涝上限
        ('soil_temperature','°C',None,None),
    ],
    'evaporation': [
        ('evaporation','mm',None,None),
        ('temperature','°C',35,40),
        ('wind_speed','m/s',None,None),
    ],
}

def _generate_site_data(site, db, now):
    """为单个站点生成传感器数据"""
    sid = site['id']; stype = site['type']
    river = site['river'] or ''
    th = RIVER_THRESHOLDS.get(river, RIVER_THRESHOLDS[''])
    base_wl = th['base']

    if stype == 'rainfall':
        # 降雨量：有时无雨，有雨时0.5-25mm
        is_rainy = random.random() < 0.35
        precip = round(random.uniform(0.5, 25) if is_rainy else 0, 1)
        cum = get_site_trend(sid,'cum',random.uniform(20,80),5,0,300)
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,threshold_high,threshold_critical,recorded_at) VALUES (?,?,?,?,?,?,?)",
            (sid,'precipitation',precip,'mm/h',20,50,now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,threshold_high,recorded_at) VALUES (?,?,?,?,?,?)",
            (sid,'cumulative_rainfall',cum,'mm',100,now))
        if precip > 50:
            create_alert_internal(db,sid,'precipitation',precip,'red',f'小时雨量超危急！{precip}mm')
        elif precip > 20:
            create_alert_internal(db,sid,'precipitation',precip,'yellow',f'小时雨量达警戒 {precip}mm')

    elif stype == 'water_level':
        wl = get_site_trend(sid,'wl',base_wl,0.06,base_wl-2,th['critical']+1)
        flow = get_site_trend(sid,'flow',round(random.uniform(200,2000),0),50,10,5000)
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,threshold_high,threshold_critical,recorded_at) VALUES (?,?,?,?,?,?,?)",
            (sid,'water_level',wl,'m',th['high'],th['critical'],now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'flow',flow,'m³/s',now))
        if wl > th['critical']:
            create_alert_internal(db,sid,'water_level',wl,'red',f'{river}水位超危急！{wl}m > {th["critical"]}m')
        elif wl > th['high']:
            create_alert_internal(db,sid,'water_level',wl,'yellow',f'{river}水位超警戒 {wl}m > {th["high"]}m')

    elif stype == 'hydrology':
        wl = get_site_trend(sid,'wl_h',base_wl,0.05,base_wl-1,th['critical']+0.5)
        vel = get_site_trend(sid,'vel',2.5,0.15,0.3,6.0)
        flow = get_site_trend(sid,'flow_h',round(random.uniform(300,3000),0),80,20,8000)
        precip = round(random.uniform(0, 15) if random.random() < 0.3 else 0, 1)
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,threshold_high,threshold_critical,recorded_at) VALUES (?,?,?,?,?,?,?)",
            (sid,'water_level',wl,'m',th['high'],th['critical'],now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'velocity',vel,'m/s',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'flow',flow,'m³/s',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,threshold_high,threshold_critical,recorded_at) VALUES (?,?,?,?,?,?,?)",
            (sid,'precipitation',precip,'mm/h',20,50,now))
        if wl > th['critical']:
            create_alert_internal(db,sid,'water_level',wl,'red',f'{site["name"]}水位超危急！{wl}m')
        elif wl > th['high']:
            create_alert_internal(db,sid,'water_level',wl,'yellow',f'{site["name"]}水位超警戒 {wl}m')

    elif stype == 'soil_moisture':
        sm = get_site_trend(sid,'sm',55,1.5,15,100)
        st = get_site_trend(sid,'st',22,0.5,5,45)
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,threshold_high,threshold_critical,recorded_at) VALUES (?,?,?,?,?,?,?)",
            (sid,'soil_moisture',sm,'%',90,None,now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'soil_temperature',st,'°C',now))
        if sm > 90:
            create_alert_internal(db,sid,'soil_moisture',sm,'yellow',f'土壤含水量过高 {sm}%（渍涝风险）')
        elif sm < 20:
            create_alert_internal(db,sid,'soil_moisture',sm,'yellow',f'土壤含水量过低 {sm}%（干旱风险）')

    elif stype == 'evaporation':
        evap = get_site_trend(sid,'evap',4.0,0.3,0,15)
        temp = get_site_trend(sid,'temp_e',28,1.5,10,45)
        wind = get_site_trend(sid,'wind_e',3.0,0.5,0,12)
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'evaporation',evap,'mm',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,threshold_high,recorded_at) VALUES (?,?,?,?,?,?)",
            (sid,'temperature',temp,'°C',40,now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'wind_speed',wind,'m/s',now))
        if temp > 40:
            create_alert_internal(db,sid,'temperature',temp,'orange',f'极端高温 {temp}°C')
        elif temp > 35:
            create_alert_internal(db,sid,'temperature',temp,'yellow',f'高温预警 {temp}°C')

    elif stype == 'groundwater':
        gwl = get_site_trend(sid,'gwl',25,0.5,5,50)
        wq = get_site_trend(sid,'wq',7.0,0.15,5.5,9.0)
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'groundwater_level',gwl,'m',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,threshold_high,recorded_at) VALUES (?,?,?,?,?,?)",
            (sid,'water_quality',wq,'pH',8.5,now))
        if wq > 8.5:
            create_alert_internal(db,sid,'water_quality',wq,'yellow',f'地下水水质异常 pH{wq}')

    elif stype == 'station_yard':
        temp_s = get_site_trend(sid,'temp_s',26,1.0,10,45)
        noise = get_site_trend(sid,'noise',55,2,30,90)
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'temperature',temp_s,'°C',now))
        db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",
            (sid,'noise',noise,'dB',now))
        if noise > 80:
            create_alert_internal(db,sid,'noise',noise,'yellow',f'站院噪音超标 {noise}dB')

def _scheduler_db():
    """专用调度器数据库连接（超时5秒，异步写入，避免阻塞API）"""
    db = sqlite3.connect(DB_PATH, timeout=5, check_same_thread=False)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA busy_timeout=5000")
    db.execute("PRAGMA synchronous=OFF")
    return db

def generate_sensor_data():
    """每30秒生成模拟传感器数据"""
    PRESET_OFFLINE = {5, 108, 193}  # 预设离线站点，跳过状态更新
    db = None
    try:
        db = _scheduler_db()
    except Exception as e:
        print(f'[Sim] 调度器连接失败: {e}')
        return
    try:
        sites = db.execute("SELECT * FROM sites").fetchall()
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        for idx, site in enumerate(sites):
            sid = site['id']
            try:
                _generate_site_data(site, db, now)
            except Exception as e:
                if 'database is locked' in str(e):
                    print(f'[Sim] DB locked, skip site {sid}')
                else:
                    print(f'[Sim] site {sid} error: {e}')

            # 设备在线/离线状态切换
            if sid in PRESET_OFFLINE: continue
            new_status = 'online'
            db.execute("UPDATE device_shadows SET status=?, last_data_time=? WHERE site_id=?",
                       (new_status, now if new_status == 'online' else None, sid))
            db.execute("UPDATE sites SET status=?, last_heartbeat=? WHERE id=?",
                       (new_status, now if new_status == 'online' else None, sid))

            # 每个站点单独提交，释放写锁，让API请求能快速插入
            try:
                db.commit()
            except Exception as e:
                print(f'[Sim] commit fail site {sid}: {e}')

        # 更新设备时间戳
        db.execute("UPDATE device_shadows SET last_data_time=? WHERE status='online'", (now,))

        # === 数据到报率模拟（在同一个连接中完成） ===
        today = datetime.now().strftime('%Y-%m-%d')
        for site in sites:
            metrics_map = {
                'rainfall': 'rainfall',
                'water_level': 'water_level',
                'hydrology': 'water_level',
                'soil_moisture': 'soil_moisture',
                'evaporation': 'evaporation',
                'groundwater': 'water_level',
                'station_yard': 'environment',
            }
            m = metrics_map.get(site['type'])
            if not m: continue
            is_miss = random.random() < 0.08
            existing = db.execute(
                "SELECT id, expected_count, actual_count FROM data_arrival WHERE site_id=? AND date=? AND metric=?",
                (site['id'], today, m)
            ).fetchone()
            if existing:
                exp = existing['expected_count'] + 1
                act = existing['actual_count'] + (0 if is_miss else 1)
                rate = round(act / exp * 100, 1)
                db.execute("UPDATE data_arrival SET expected_count=?, actual_count=?, arrival_rate=? WHERE id=?",
                           (exp, act, rate, existing['id']))
            else:
                db.execute("INSERT INTO data_arrival (site_id,date,metric,expected_count,actual_count,arrival_rate) VALUES (?,?,?,1,?,?)",
                           (site['id'], today, m, 0 if is_miss else 1, 100 if not is_miss else 0))

        # === 天气数据 ===
        # 启动时尝试获取实时天气（不插入模拟数据，让 API 请求时自动刷新）
        fetch_real_weather()
        
        # === 离线设备告警 ===
        offline_devices = db.execute("""
            SELECT d.site_id, d.device_name, d.device_code, s.name as site_name
            FROM device_shadows d JOIN sites s ON d.site_id=s.id
            WHERE d.status='offline'
        """).fetchall()
        for dev in offline_devices:
            create_alert_internal(db, dev['site_id'], 'device_status', 0, 'yellow',
                f"设备离线: {dev['device_name']} ({dev['device_code']}) · {dev['site_name']}")
    except Exception as e:
        print(f'[Sim] 数据生成异常: {e}')
    finally:
        if db:
            try:
                db.close()
            except:
                pass

def create_alert_internal(db, site_id, metric, value, level, message):
    # 检查最近1小时内是否有相同的site+metric告警（含已办结），有则跳过
    exists = db.execute(
        "SELECT id FROM alerts WHERE site_id=? AND metric=? AND level=? AND created_at > datetime('now','-60 minutes')",
        (site_id, metric, level)
    ).fetchone()
    if not exists:
        db.execute(
            "INSERT INTO alerts (site_id,metric,value,level,message) VALUES (?,?,?,?,?)",
            (site_id, metric, value, level, message)
        )

# ===================== 登录认证系统 =====================

# Token存储
_tokens = {}

def _hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def seed_users():
    with get_db() as db:
        cnt = db.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if cnt > 0:
            print("[Auth] 用户已存在，跳过")
            return
        users = [
            ('admin', 'admin123', 'admin', '系统管理员', '13800000000'),
            ('zhangsan', 'yw123456', 'operator', '张建国', '13800000001'),
            ('lisi', 'yw123456', 'operator', '黎明', '13800000002'),
            ('wangwu', 'yw123456', 'operator', '王刚', '13800000003'),
            ('zhaoliu', 'yw123456', 'operator', '赵洪', '13800000004'),
        ]
        for u in users:
            db.execute("INSERT INTO users (username,password_hash,role,real_name,phone) VALUES (?,?,?,?,?)",
                       (u[0], _hash_pw(u[1]), u[2], u[3], u[4]))
        print("[Auth] 5个用户已创建")
        all_ids = [r['id'] for r in db.execute("SELECT id FROM sites").fetchall()]
        for sid in all_ids:
            db.execute("INSERT OR IGNORE INTO user_sites (user_id,site_id) VALUES (?,?)", (1, sid))
        assignments = [(2, 1, 70), (3, 71, 140), (4, 141, 210), (5, 211, 267)]
        for uid, start_id, end_id in assignments:
            for sid in range(start_id, end_id + 1):
                if sid <= max(all_ids):
                    db.execute("INSERT OR IGNORE INTO user_sites (user_id,site_id) VALUES (?,?)", (uid, sid))
        db.commit()
        print("[Auth] 站点分配完成")

def login_required(f):
    """认证中间件：从Authorization头中提取token，注入g.current_user和g.user_sites"""
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.headers.get('Authorization', '')
        token = auth.replace('Bearer ', '').strip() if auth.startswith('Bearer ') else ''
        if not token or token not in _tokens:
            return jsonify({'error': '未登录或登录已过期', 'code': 'AUTH_REQUIRED'}), 401
        user = _tokens[token]
        g.current_user = user
        with get_db() as db:
            rows = db.execute("SELECT site_id FROM user_sites WHERE user_id=?", (user['id'],)).fetchall()
        g.user_site_ids = [r['site_id'] for r in rows]
        return f(*args, **kwargs)
    return wrapper

def _filter_site_ids():
    """返回当前用户可见的site_id列表（管理员返回None=全部）"""
    site_ids = getattr(g, 'user_site_ids', None)
    if site_ids is None:
        return None
    return site_ids


# ===================== API Routes =====================

@app.route('/api/health')
def health():
    return jsonify({'status': 'ok', 'time': datetime.now().isoformat()})

# --- Sites ---
@app.route('/api/sites')
@login_required
def get_sites():
    site_ids = _filter_site_ids()
    with get_db() as db:
        if site_ids is not None:
            placeholders = ','.join('?' * len(site_ids))
            rows = db.execute(f"""
                SELECT s.id, s.code, s.name, s.type, s.lat, s.lng, s.district, s.river,
                       s.manager, s.phone, s.last_heartbeat, s.created_at,
                       COUNT(d.id) as device_count,
                       SUM(CASE WHEN d.status='offline' THEN 1 ELSE 0 END) as offline_count,
                       CASE WHEN SUM(CASE WHEN d.status='offline' THEN 1 ELSE 0 END) > 0 THEN 'offline' ELSE 'online' END as status
                FROM sites s LEFT JOIN device_shadows d ON s.id=d.site_id
                WHERE s.id IN ({placeholders})
                GROUP BY s.id ORDER BY s.id
            """, site_ids).fetchall()
        else:
            rows = db.execute("""
                SELECT s.id, s.code, s.name, s.type, s.lat, s.lng, s.district, s.river,
                       s.manager, s.phone, s.last_heartbeat, s.created_at,
                       COUNT(d.id) as device_count,
                       SUM(CASE WHEN d.status='offline' THEN 1 ELSE 0 END) as offline_count,
                       CASE WHEN SUM(CASE WHEN d.status='offline' THEN 1 ELSE 0 END) > 0 THEN 'offline' ELSE 'online' END as status
                FROM sites s LEFT JOIN device_shadows d ON s.id=d.site_id
                GROUP BY s.id ORDER BY s.id
            """).fetchall()
        result = []
        for r in rows:
            d = dict(r)
            d['status'] = r['status']
            result.append(d)
        return jsonify(result)

@app.route('/api/sites/<int:site_id>')
def get_site(site_id):
    with get_db() as db:
        site = db.execute("SELECT * FROM sites WHERE id=?", (site_id,)).fetchone()
        if not site:
            return jsonify({'error': 'not found'}), 404
        devices = db.execute("SELECT * FROM device_shadows WHERE site_id=?", (site_id,)).fetchall()
        alerts_count = db.execute("SELECT COUNT(*) as c FROM alerts WHERE site_id=? AND status='pending'", (site_id,)).fetchone()['c']
        orders_count = db.execute("SELECT COUNT(*) as c FROM work_orders WHERE site_id=? AND status NOT IN ('closed')", (site_id,)).fetchone()['c']
        site_dict = dict(site)
        # Calculate status from devices, not from sites.status
        offline_devices = [d for d in devices if d['status'] == 'offline']
        site_dict['status'] = 'offline' if len(offline_devices) > 0 else 'online'
        site_dict['devices'] = [dict(d) for d in devices]
        site_dict['active_alerts'] = alerts_count
        site_dict['open_orders'] = orders_count
        # 水库额外信息
        reservoir_extra = {
            1: {'capacity': 1280, 'flood_level': 49.5, 'critical_level': 51.5, 'normal_level': 48.0},
            2: {'capacity': 860, 'flood_level': 48.0, 'critical_level': 50.0, 'normal_level': 47.0},
        }
        if site['type'] == 'reservoir':
            extra = reservoir_extra.get(site['id'], {})
            site_dict['capacity'] = extra.get('capacity')
            site_dict['flood_level'] = extra.get('flood_level')
            site_dict['critical_level'] = extra.get('critical_level')
            site_dict['normal_level'] = extra.get('normal_level')
        return jsonify(site_dict)

# --- Sensor Data ---
@app.route('/api/data/realtime')
@login_required
def realtime_data():
    """各站点最新一条数据（优化：一次查询，不用N+1）"""
    site_ids = _filter_site_ids()
    with get_db() as db:
        # 一次查询获取所有站点的最新传感器数据
        latest = {}
        try:
            latest_rows = db.execute("""
                SELECT sd.site_id, sd.metric, sd.value, sd.unit, sd.recorded_at
                FROM sensor_data sd
                INNER JOIN (
                    SELECT site_id, MAX(recorded_at) as max_t
                    FROM sensor_data
                    GROUP BY site_id
                ) lm ON sd.site_id = lm.site_id AND sd.recorded_at = lm.max_t
            """).fetchall()
            for r in latest_rows:
                latest[r['site_id']] = r
        except:
            pass
        site_sql = """SELECT s.id, s.code, s.name, s.type, s.lat, s.lng,
                   CASE WHEN COUNT(d.id) = 0 THEN 'online'
                        WHEN SUM(CASE WHEN d.status='offline' THEN 1 ELSE 0 END) > 0 THEN 'offline'
                        ELSE 'online' END as status
            FROM sites s LEFT JOIN device_shadows d ON s.id=d.site_id"""
        site_params = []
        if site_ids is not None:
            placeholders = ','.join('?' * len(site_ids))
            site_sql += f" WHERE s.id IN ({placeholders})"
            site_params = site_ids
        site_sql += " GROUP BY s.id"
        sites = db.execute(site_sql, site_params).fetchall()
        result = []
        for s in sites:
            row = latest.get(s['id'])
            site_dict = dict(s)
            site_dict['latest_value'] = round(row['value'],2) if row else 0
            site_dict['latest_metric'] = row['metric'] if row else ''
            site_dict['latest_unit'] = row['unit'] if row else ''
            site_dict['latest_time'] = row['recorded_at'] if row else ''
            result.append(site_dict)
        return jsonify(result)

@app.route('/api/data/site/<int:site_id>')
def site_data(site_id):
    """站点最近2小时数据"""
    limit = request.args.get('limit', 50, type=int)
    with get_db() as db:
        rows = db.execute(
            "SELECT metric, value, unit, recorded_at FROM sensor_data WHERE site_id=? ORDER BY recorded_at DESC LIMIT ?",
            (site_id, limit)
        ).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/data/overview')
@login_required
def data_overview():
    site_ids = _filter_site_ids()
    with get_db() as db:
        if site_ids is not None:
            placeholders = ','.join('?' * len(site_ids))
            total_sites = db.execute(f"SELECT COUNT(*) as c FROM sites WHERE id IN ({placeholders})", site_ids).fetchone()['c']
            online_sites = db.execute(f"SELECT COUNT(*) as c FROM sites WHERE status='online' AND id IN ({placeholders})", site_ids).fetchone()['c']
            device_total = db.execute(f"SELECT COUNT(*) as c FROM device_shadows WHERE site_id IN ({placeholders})", site_ids).fetchone()['c']
            device_online = db.execute(f"SELECT COUNT(*) as c FROM device_shadows WHERE status='online' AND site_id IN ({placeholders})", site_ids).fetchone()['c']
            active_alerts = db.execute(f"SELECT COUNT(*) as c FROM alerts WHERE status='pending' AND site_id IN ({placeholders})", site_ids).fetchone()['c']
            open_orders = db.execute(f"SELECT COUNT(*) as c FROM work_orders WHERE status NOT IN ('closed') AND site_id IN ({placeholders})", site_ids).fetchone()['c']
        else:
            total_sites = db.execute("SELECT COUNT(*) as c FROM sites").fetchone()['c']
            online_sites = db.execute("SELECT COUNT(*) as c FROM sites WHERE status='online'").fetchone()['c']
            device_total = db.execute("SELECT COUNT(*) as c FROM device_shadows").fetchone()['c']
            device_online = db.execute("SELECT COUNT(*) as c FROM device_shadows WHERE status='online'").fetchone()['c']
            active_alerts = db.execute("SELECT COUNT(*) as c FROM alerts WHERE status='pending'").fetchone()['c']
            open_orders = db.execute("SELECT COUNT(*) as c FROM work_orders WHERE status NOT IN ('closed')").fetchone()['c']
        return jsonify({
            'total_sites': total_sites, 'online_sites': online_sites,
            'device_total': device_total, 'device_online': device_online,
            'active_alerts': active_alerts, 'open_orders': open_orders
        })

# --- Alerts ---
@app.route('/api/alerts')
@login_required
def get_alerts():
    site_ids = _filter_site_ids()
    status = request.args.get('status', '')
    limit = request.args.get('limit', 50, type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    with get_db() as db:
        q = """
            SELECT a.*, s.name as site_name, s.code as site_code
            FROM alerts a LEFT JOIN sites s ON a.site_id=s.id
            WHERE 1=1
        """
        params = []
        if site_ids is not None:
            placeholders = ','.join('?' * len(site_ids))
            q += f" AND a.site_id IN ({placeholders})"
            params.extend(site_ids)
        if status:
            q += " AND a.status=?"
            params.append(status)
        if date_from:
            q += " AND a.created_at>=?"
            params.append(date_from)
        if date_to:
            q += " AND a.created_at<=?"
            params.append(date_to + ' 23:59:59')
        q += " ORDER BY CASE a.level WHEN 'red' THEN 1 WHEN 'orange' THEN 2 WHEN 'yellow' THEN 3 ELSE 4 END, a.created_at DESC LIMIT ?"
        params.append(limit)
        return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/alerts/<int:alert_id>/acknowledge', methods=['POST'])
def acknowledge_alert(alert_id):
    data = request.json or {}
    operator = data.get('operator', '系统')
    with get_db() as db:
        db.execute("UPDATE alerts SET status='acknowledged' WHERE id=?", (alert_id,))
        # 记录时间线
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'acknowledged', operator, '确认告警'))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/alerts/<int:alert_id>/resolve', methods=['POST'])
def resolve_alert(alert_id):
    """办结告警，支持办结原因（reason）"""
    data = request.json or {}
    operator = data.get('operator', '系统')
    reason = data.get('reason', '办结告警')
    remark = data.get('remark', '')
    # reason可选值: 误报 / 仪器正常偏差 / 已自动恢复 / 已人工处理 / 自定义
    full_remark = reason + (' - ' + remark if remark else '')
    with get_db() as db:
        db.execute("UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime') WHERE id=?", (alert_id,))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'resolved', operator, full_remark))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/alerts/<int:alert_id>/ack-resolve', methods=['POST'])
def ack_resolve_alert(alert_id):
    """一键确认并办结（跳过已确认状态，直接pending→resolved）"""
    data = request.json or {}
    operator = data.get('operator', '系统')
    remark = data.get('remark', '一键办结')
    with get_db() as db:
        db.execute("UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime') WHERE id=? AND status='pending'", (alert_id,))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'acknowledged', operator, '确认告警'))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'resolved', operator, remark))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/alerts/<int:alert_id>/urge', methods=['POST'])
def urge_alert(alert_id):
    """告警督办，支持时限和协办单位"""
    data = request.json or {}
    operator = data.get('operator', '系统')
    remark = data.get('remark', '督办告警')
    deadline = data.get('deadline', '')
    cooperator = data.get('cooperator', '')
    # 将额外信息拼入remark
    extra = []
    if deadline: extra.append('限办:'+deadline)
    if cooperator: extra.append('协办:'+cooperator)
    full_remark = remark + (' | ' + '; '.join(extra) if extra else '')
    # 更新数据库中的response_deadline字段
    with get_db() as db:
        db.execute("UPDATE alerts SET urge_count=COALESCE(urge_count,0)+1, last_urged_at=datetime('now','localtime') WHERE id=?", (alert_id,))
        if deadline:
            db.execute("UPDATE alerts SET response_deadline=? WHERE id=?", (deadline, alert_id))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'urged', operator, full_remark))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/alerts/<int:alert_id>/undo-acknowledge', methods=['POST'])
def undo_acknowledge_alert(alert_id):
    """撤销告警确认，将状态改回pending"""
    data = request.json or {}
    operator = data.get('operator', '系统')
    remark = data.get('remark', '撤销确认')
    with get_db() as db:
        db.execute("UPDATE alerts SET status='pending', resolved_at=NULL WHERE id=?", (alert_id,))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'undo_acknowledge', operator, remark))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/alerts/<int:alert_id>/convert-order', methods=['POST'])
def convert_alert_to_order(alert_id):
    """告警转工单"""
    data = request.json or {}
    operator = data.get('operator', '系统')
    with get_db() as db:
        alert = db.execute("SELECT * FROM alerts WHERE id=?", (alert_id,)).fetchone()
        if not alert:
            return jsonify({'error': 'not found'}), 404
        now = datetime.now()
        order_no = f"WO-{now.strftime('%Y%m%d')}-{random.randint(100,999)}"
        level = data.get('level', alert['level'])
        if level == 'red':
            order_level = 'critical'
        elif level == 'orange':
            order_level = 'urgent'
        else:
            order_level = 'normal'
        sla_hours = {'normal': 72, 'urgent': 24, 'critical': 2}.get(order_level, 72)
        sla_deadline = (now + timedelta(hours=sla_hours)).strftime('%Y-%m-%d %H:%M')
        assignee = data.get('assignee', '')
        db.execute("""
            INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,assignee,status,sla_deadline)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            order_no, alert['site_id'], 'auto', '告警转工单',
            order_level, f"[告警转] {alert['message']}", alert['message'],
            assignee, 'pending', sla_deadline
        ))
        # 更新告警关联工单号 + 状态改为已确认
        db.execute("UPDATE alerts SET related_order_no=?, status='acknowledged' WHERE id=?", (order_no, alert_id))
        # 记录时间线
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('alert', alert_id, 'converted', operator, f'转工单 {order_no}'))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('order', 0, 'created', operator, f'告警{alert_id}转工单-{order_no}'))
        db.commit()
        return jsonify({'success': True, 'order_no': order_no})

@app.route('/api/alerts/batch', methods=['POST'])
def batch_alert_operations():
    """告警批量操作: acknowledge/resolve/urge/convert"""
    data = request.json or {}
    ids = data.get('ids', [])
    action = data.get('action', '')
    operator = data.get('operator', '系统')
    if not ids or not action:
        return jsonify({'error': 'ids and action required'}), 400
    with get_db() as db:
        if action == 'acknowledge':
            placeholders = ','.join(['?'] * len(ids))
            db.execute(f"UPDATE alerts SET status='acknowledged' WHERE id IN ({placeholders})", ids)
            for aid in ids:
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('alert', aid, 'acknowledged', operator, '批量确认'))
        elif action == 'resolve':
            reason = data.get('reason', '批量办结')
            placeholders = ','.join(['?'] * len(ids))
            db.execute(f"UPDATE alerts SET status='resolved', resolved_at=datetime('now','localtime') WHERE id IN ({placeholders}) AND status='pending'", ids)
            for aid in ids:
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('alert', aid, 'resolved', operator, reason))
        elif action == 'urge':
            remark = data.get('remark', '批量督办')
            for aid in ids:
                db.execute("UPDATE alerts SET urge_count=COALESCE(urge_count,0)+1, last_urged_at=datetime('now','localtime') WHERE id=?", (aid,))
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('alert', aid, 'urged', operator, remark))
        elif action == 'convert':
            for aid in ids:
                alert = db.execute("SELECT * FROM alerts WHERE id=?", (aid,)).fetchone()
                if not alert:
                    continue
                now = datetime.now()
                order_no = f"WO-{now.strftime('%Y%m%d')}-{random.randint(100,999)}"
                level = alert['level']
                order_level = 'critical' if level == 'red' else ('urgent' if level == 'orange' else 'normal')
                sla_hours = {'normal': 72, 'urgent': 24, 'critical': 2}.get(order_level, 72)
                sla_deadline = (now + timedelta(hours=sla_hours)).strftime('%Y-%m-%d %H:%M')
                db.execute("""
                    INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,status,sla_deadline)
                    VALUES (?,?,?,?,?,?,?,?,?)
                """, (order_no, alert['site_id'], 'auto', '告警批量转工单', order_level,
                      f"[告警转] {alert['message']}", alert['message'], 'pending', sla_deadline))
                db.execute("UPDATE alerts SET related_order_no=?, status='acknowledged' WHERE id=?", (order_no, aid))
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('alert', aid, 'converted', operator, f'批量转工单 {order_no}'))
        else:
            return jsonify({'error': f'unknown action: {action}'}), 400
        db.commit()
        return jsonify({'success': True, 'count': len(ids)})

@app.route('/api/timeline')
def get_timeline():
    """时间线查询，可按来源过滤"""
    source_type = request.args.get('source_type', '')
    source_id = request.args.get('source_id', '', type=int) if request.args.get('source_id') else None
    limit = request.args.get('limit', 50, type=int)
    with get_db() as db:
        q = "SELECT * FROM timeline_events WHERE 1=1"
        params = []
        if source_type:
            q += " AND source_type=?"
            params.append(source_type)
        if source_id is not None:
            q += " AND source_id=?"
            params.append(source_id)
        q += " ORDER BY created_at DESC LIMIT ?"
        params.append(limit)
        return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/alerts/statistics')
@login_required
def alert_statistics():
    site_ids = _filter_site_ids()
    status = request.args.get('status', '')
    status_where = ''
    params = []
    if status:
        status_where = ' WHERE status=?'
        params.append(status)
    with get_db() as db:
        total = db.execute(f"SELECT COUNT(*) as c FROM alerts{status_where}", params).fetchone()['c']
        by_level = {}
        for lv in ['red','orange','yellow','blue']:
            lv_params = params + [lv]
            by_level[lv] = db.execute(f"SELECT COUNT(*) as c FROM alerts{status_where + ' AND level=?' if status else ' WHERE level=?'}", lv_params).fetchone()['c']
        by_status = {}
        if not status:
            for st in ['pending','acknowledged','resolved']:
                by_status[st] = db.execute("SELECT COUNT(*) as c FROM alerts WHERE status=?",(st,)).fetchone()['c']
        else:
            by_status[status] = total
        return jsonify({'total':total, 'by_level':by_level, 'by_status':by_status})

# --- Work Orders ---
@app.route('/api/workorders')
@login_required
def get_workorders():
    status = request.args.get('status', '')
    limit = request.args.get('limit', 50, type=int)
    site_ids = _filter_site_ids()
    with get_db() as db:
        q = """
            SELECT w.*, s.name as site_name
            FROM work_orders w LEFT JOIN sites s ON w.site_id=s.id
            WHERE 1=1
        """
        params = []
        if site_ids is not None:
            ph = ','.join('?' * len(site_ids))
            q += f" AND w.site_id IN ({ph})"
            params.extend(site_ids)
        if status:
            q += " AND w.status=?"
            params.append(status)
        q += " ORDER BY w.created_at DESC LIMIT ?"
        params.append(limit)
        return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/workorders', methods=['POST'])
def create_workorder():
    data = request.json
    with get_db() as db:
        now = datetime.now()
        order_no = f"WO-{now.strftime('%Y%m%d')}-{random.randint(100,999)}"
        sla_hours = {'normal': 72, 'urgent': 24, 'critical': 2}.get(data.get('level','normal'), 72)
        sla_deadline = (now + timedelta(hours=sla_hours)).strftime('%Y-%m-%d %H:%M')
        db.execute("""
            INSERT INTO work_orders (order_no,site_id,source,event_type,level,title,description,images,assignee,status,sla_deadline)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (
            order_no, data.get('site_id'), data.get('source','manual'),
            data.get('event_type',''), data.get('level','normal'),
            data.get('title',''), data.get('description',''),
            data.get('images',''), data.get('assignee',''),
            'pending', sla_deadline
        ))
        db.commit()
        return jsonify({'success': True, 'order_no': order_no})

@app.route('/api/workorders/<order_no>/status', methods=['PUT'])
def update_workorder_status(order_no):
    data = request.json
    new_status = data.get('status')
    valid_transitions = {
        'pending': ['accepted'],
        'accepted': ['in_progress'],
        'in_progress': ['reviewing', 'accepted'],
        'reviewing': ['acceptance', 'in_progress'],
        'acceptance': ['closed'],
    }
    with get_db() as db:
        cur = db.execute("SELECT status FROM work_orders WHERE order_no=?", (order_no,)).fetchone()
        if not cur:
            return jsonify({'error': 'not found'}), 404
        if new_status not in valid_transitions.get(cur['status'], []):
            return jsonify({'error': f'invalid transition from {cur["status"]} to {new_status}'}), 400
        updates = ["status=?"]
        params = [new_status]
        if new_status == 'closed':
            updates.append("resolved_at=datetime('now','localtime')")
        if 'remark' in data:
            updates.append("remark=?")
            params.append(data['remark'])
        if 'satisfaction' in data:
            updates.append("satisfaction=?")
            params.append(data['satisfaction'])
        params.append(order_no)
        db.execute(f"UPDATE work_orders SET {','.join(updates)} WHERE order_no=?", params)
        # 时间线记录
        operator = data.get('operator', '系统')
        status_cn = {'pending':'待受理','accepted':'已受理','generated':'已生成','dispatched':'已派发','in_progress':'处置中','reviewing':'审核中','acceptance':'验收中','closed':'已关闭'}
        event_label = status_cn.get(new_status, new_status)
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('order', 0, new_status, operator, f'工单{order_no} → {event_label}'))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/workorders/<orderNo>/photos')
@login_required
def workorder_photos(orderNo):
    """返回工单关联的巡检照片（通过站点+检查项描述匹配）"""
    with get_db() as db:
        wo = db.execute("SELECT * FROM work_orders WHERE order_no=?", (orderNo,)).fetchone()
        if not wo:
            return jsonify({'error': '工单不存在'}), 404
        # 从描述中提取可能的检查项名称
        desc = wo.get('description', '')
        check_items = []
        if '检查项: ' in desc:
            check_items.append(desc.split('检查项: ')[-1].strip())
        # 按站点匹配已完成的巡检任务（有照片的）
        photos = []
        if wo.get('site_id'):
            rows = db.execute(
                "SELECT check_item, photo, remark, check_time FROM inspection_tasks WHERE site_id=? AND photo IS NOT NULL AND photo != '' ORDER BY check_time DESC",
                (wo['site_id'],)
            ).fetchall()
            for r in rows:
                photos.append({
                    'check_item': r['check_item'],
                    'photo': r['photo'],
                    'remark': r['remark'] or '',
                    'time': r['check_time'] or '',
                })
        return jsonify({'success': True, 'photos': photos})

@app.route('/api/workorders/statistics')
@login_required
def workorder_statistics():
    site_ids = _filter_site_ids()
    with get_db() as db:
        if site_ids is not None:
            ph = ','.join('?' * len(site_ids))
            total = db.execute(f"SELECT COUNT(*) as c FROM work_orders WHERE site_id IN ({ph})", site_ids).fetchone()['c']
            by_status = {}
            for st in ['pending','accepted','generated','dispatched','in_progress','reviewing','acceptance','closed']:
                by_status[st] = db.execute(f"SELECT COUNT(*) as c FROM work_orders WHERE status=? AND site_id IN ({ph})", [st] + site_ids).fetchone()['c']
            today = datetime.now().strftime('%Y-%m-%d')
            today_new = db.execute(f"SELECT COUNT(*) as c FROM work_orders WHERE date(created_at)=? AND site_id IN ({ph})", [today] + site_ids).fetchone()['c']
            today_closed = db.execute(f"SELECT COUNT(*) as c FROM work_orders WHERE date(resolved_at)=? AND site_id IN ({ph})", [today] + site_ids).fetchone()['c']
        else:
            total = db.execute("SELECT COUNT(*) as c FROM work_orders").fetchone()['c']
            by_status = {}
            for st in ['pending','accepted','generated','dispatched','in_progress','reviewing','acceptance','closed']:
                by_status[st] = db.execute("SELECT COUNT(*) as c FROM work_orders WHERE status=?",(st,)).fetchone()['c']
            today = datetime.now().strftime('%Y-%m-%d')
            today_new = db.execute("SELECT COUNT(*) as c FROM work_orders WHERE date(created_at)=?",(today,)).fetchone()['c']
            today_closed = db.execute("SELECT COUNT(*) as c FROM work_orders WHERE date(resolved_at)=?",(today,)).fetchone()['c']
        return jsonify({'total':total, 'by_status':by_status, 'today_new':today_new, 'today_closed':today_closed})

# --- Inspections ---
@app.route('/api/inspections')
@login_required
def get_inspections():
    site_ids = _filter_site_ids()
    with get_db() as db:
        q = """
            SELECT p.*, s.name as site_name, s.code as site_code,
                (SELECT COUNT(*) FROM inspection_tasks t WHERE t.plan_id=p.id) as total_items,
                (SELECT COUNT(*) FROM inspection_tasks t WHERE t.plan_id=p.id AND t.result IS NOT NULL) as completed_items
            FROM inspection_plans p LEFT JOIN sites s ON p.site_id=s.id
            WHERE 1=1
        """
        params = []
        if site_ids is not None:
            ph = ','.join('?' * len(site_ids))
            q += f" AND p.site_id IN ({ph})"
            params.extend(site_ids)
        q += " ORDER BY p.created_at DESC"
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/inspections', methods=['POST'])
def create_inspection():
    data = request.json
    with get_db() as db:
        scheme_id = data.get('scheme_id')
        cursor = db.execute("""
            INSERT INTO inspection_plans (plan_name,site_id,type,start_date,end_date,period,description,scheme_id)
            VALUES (?,?,?,?,?,?,?,?)
        """, (data['plan_name'],data['site_id'],data['type'],data['start_date'],data['end_date'],data.get('period','once'),data.get('description',''),scheme_id))
        plan_id = cursor.lastrowid
        # 生成检查项：优先从scheme_id加载，否则用check_items
        check_items = data.get('check_items', [])
        if scheme_id:
            scheme_items = db.execute("SELECT check_item FROM inspection_scheme_items WHERE scheme_id=? ORDER BY sort_order",(scheme_id,)).fetchall()
            if scheme_items:
                check_items = [r['check_item'] for r in scheme_items]
        if not check_items:
            check_items = ['坝体外观检查','溢洪道检查','放水设施检查','监测设备检查','防汛物资检查','管理设施检查']
        for item in check_items:
            db.execute(
                "INSERT INTO inspection_tasks (plan_id,site_id,check_item) VALUES (?,?,?)",
                (plan_id, data['site_id'], item)
            )
        db.commit()
        # 时间线记录
        operator = data.get('operator', '系统')
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('inspection', plan_id, 'created', operator, f'创建巡检计划-{data["plan_name"]}'))
        db.commit()
        return jsonify({'success': True, 'plan_id': plan_id})

@app.route('/api/inspections/<int:plan_id>/tasks')
def get_inspection_tasks(plan_id):
    with get_db() as db:
        rows = db.execute("SELECT * FROM inspection_tasks WHERE plan_id=? ORDER BY id", (plan_id,)).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/inspections/tasks/<int:task_id>', methods=['PUT'])
def update_inspection_task(task_id):
    data = request.json
    with get_db() as db:
        db.execute("""
            UPDATE inspection_tasks SET result=?, photo=?, gps_lat=?, gps_lng=?, check_time=?, remark=?
            WHERE id=?
        """, (data.get('result'),data.get('photo'),data.get('gps_lat'),data.get('gps_lng'),
              data.get('check_time'),data.get('remark'),task_id))
        # 更新计划状态
        task = db.execute("SELECT plan_id FROM inspection_tasks WHERE id=?", (task_id,)).fetchone()
        if task:
            incomplete = db.execute(
                "SELECT COUNT(*) as c FROM inspection_tasks WHERE plan_id=? AND result IS NULL",
                (task['plan_id'],)
            ).fetchone()['c']
            if incomplete == 0:
                db.execute("UPDATE inspection_plans SET status='completed' WHERE id=?", (task['plan_id'],))
                plan = db.execute("SELECT plan_name FROM inspection_plans WHERE id=?", (task['plan_id'],)).fetchone()
                if plan:
                    db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                               ('inspection', task['plan_id'], 'completed', '系统', f'巡检计划完成-{plan["plan_name"]}'))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/inspections/statistics')
@login_required
def inspection_statistics():
    site_ids = _filter_site_ids()
    with get_db() as db:
        if site_ids is not None:
            ph = ','.join('?' * len(site_ids))
            total_plans = db.execute(f"SELECT COUNT(*) as c FROM inspection_plans WHERE site_id IN ({ph})", site_ids).fetchone()['c']
            done = db.execute(f"SELECT COUNT(*) as c FROM inspection_plans WHERE status='completed' AND site_id IN ({ph})", site_ids).fetchone()['c']
            total_tasks = db.execute(f"SELECT COUNT(*) as c FROM inspection_tasks WHERE site_id IN ({ph})", site_ids).fetchone()['c']
            done_tasks = db.execute(f"SELECT COUNT(*) as c FROM inspection_tasks WHERE result IS NOT NULL AND site_id IN ({ph})", site_ids).fetchone()['c']
            abnormal = db.execute(f"SELECT COUNT(*) as c FROM inspection_tasks WHERE result='abnormal' AND site_id IN ({ph})", site_ids).fetchone()['c']
        else:
            total_plans = db.execute("SELECT COUNT(*) as c FROM inspection_plans").fetchone()['c']
            done = db.execute("SELECT COUNT(*) as c FROM inspection_plans WHERE status='completed'").fetchone()['c']
            total_tasks = db.execute("SELECT COUNT(*) as c FROM inspection_tasks").fetchone()['c']
            done_tasks = db.execute("SELECT COUNT(*) as c FROM inspection_tasks WHERE result IS NOT NULL").fetchone()['c']
            abnormal = db.execute("SELECT COUNT(*) as c FROM inspection_tasks WHERE result='abnormal'").fetchone()['c']
        return jsonify({
            'total_plans':total_plans, 'completed_plans':done,
            'total_tasks':total_tasks, 'completed_tasks':done_tasks,
            'abnormal_count':abnormal
        })


# ===================== 巡检方案管理 API =====================

DEFAULT_CHECK_ITEMS = {
    '水工建筑物': ['坝体/堤防外观检查','溢洪道/泄洪设施检查','放水涵洞/输水设施检查','护坡/护岸完整性检查','变形/位移观测'],
    '金属结构': ['闸门启闭机运行检查','钢丝绳/吊杆磨损检查','止水橡胶/密封件检查'],
    '监测设备': ['水位计运行状态及精度校验','雨量计清洁及校准','流量计/流速仪运行检查','土壤墒情传感器检查','蒸发皿清洁及补水'],
    '通信设备': ['RTU遥测终端运行状态','通信模块(4G/北斗)信号检测','数据采集频率/完整性校验'],
    '供电系统': ['太阳能板清洁及朝向检查','蓄电池电压/容量测试','充放电控制器检查','供电线路/防雷器检查'],
    '安全防护': ['防雷接地电阻测试','围栏/门锁/警示标识','视频监控设备检查'],
    '环境维护': ['站房/站院卫生清理','排水沟/截水沟疏通','杂草清除/防鼠防虫'],
    '应急管理': ['备品备件库存清点','应急物资/工具检查','防汛预案/操作规程上墙'],
    '自定义': []
}

DAY_ITEMS = ['水工建筑物','金属结构','监测设备','通信设备','环境维护']
WEEK_ITEMS = ['水工建筑物','金属结构','监测设备','通信设备','供电系统','安全防护','环境维护','应急管理']
MONTH_ITEMS = ['水工建筑物','金属结构','监测设备','通信设备','供电系统','安全防护','环境维护','应急管理']

@app.route('/api/schemes/template')
def download_scheme_template():
    try: import openpyxl
    except: return jsonify({'error':'openpyxl未安装'}),500
    import os
    from openpyxl.worksheet.datavalidation import DataValidation
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = '巡检方案模板'
    headers = ['站点名称','分类','检查项','日方案','周方案','月方案']
    for col, h in enumerate(headers,1):
        ws.cell(row=1,column=col,value=h); ws.cell(row=1,column=col).font = openpyxl.styles.Font(bold=True)
    # Collect all check items
    all_cats = [c for c in DEFAULT_CHECK_ITEMS if DEFAULT_CHECK_ITEMS[c]]
    all_items = []
    for cat in all_cats:
        for item in DEFAULT_CHECK_ITEMS[cat]:
            all_items.append((cat, item))
    
    row = 2
    for cat, item in all_items:
        ws.cell(row=row,column=1,value=''); ws.cell(row=row,column=2,value=cat); ws.cell(row=row,column=3,value=item)
        ws.cell(row=row,column=4,value='✓' if cat in DAY_ITEMS else '')
        ws.cell(row=row,column=5,value='✓' if cat in WEEK_ITEMS else '')
        ws.cell(row=row,column=6,value='✓' if cat in MONTH_ITEMS else '')
        row += 1
    
    # Add data validation: dropdown for 分类 (column B) and 检查项 (column C)
    # Write list sources to a hidden helper sheet for more reliable dropdowns
    hs = wb.create_sheet('_list', 0); hs.sheet_state = 'hidden'
    for idx, cat in enumerate(all_cats):
        hs.cell(row=idx+1, column=1, value=cat)
    for idx, (_, item) in enumerate(all_items):
        hs.cell(row=idx+1, column=2, value=item)
    
    dv_cat = DataValidation(type="list", formula1=f'_list!$A$1:$A${len(all_cats)}', allow_blank=True)
    dv_cat.error = '请从下拉列表中选择分类'
    dv_cat.errorTitle = '输入错误'
    ws.add_data_validation(dv_cat)
    dv_cat.add(f'B2:B{row-1}')
    
    dv_item = DataValidation(type="list", formula1=f'_list!$B$1:$B${len(all_items)}', allow_blank=True)
    dv_item.error = '请从下拉列表中选择检查项'
    dv_item.errorTitle = '输入错误'
    ws.add_data_validation(dv_item)
    dv_item.add(f'C2:C{row-1}')
    
    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'frontend', '_template.xlsx')
    wb.save(out_path)
    return send_file(out_path, as_attachment=True, download_name='巡检方案导入模板.xlsx')

@app.route('/api/schemes/import', methods=['POST'])
def import_schemes():
    """导入站点级方案表：每行=站点+检查项，匹配站点名后写入对应方案"""
    try: import openpyxl
    except: return jsonify({'error':'openpyxl未安装'}),500
    file = request.files.get('file')
    if not file: return jsonify({'error':'请上传文件'}),400
    wb = openpyxl.load_workbook(file); ws = wb.active
    def _yes(v): return str(v).strip() in ('✓','√','Y','y','1','是','yes')
    created = 0
    with get_db() as db:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row)<3: continue
            site_name = str(row[0]).strip() if row[0] else ''
            cat = str(row[1]).strip() if row[1] else ''
            item = str(row[2]).strip() if row[2] else ''
            day_flag = row[3] if len(row)>3 and row[3] else ''
            week_flag = row[4] if len(row)>4 and row[4] else ''
            month_flag = row[5] if len(row)>5 and row[5] else ''
            if not site_name or not item: continue
            site = db.execute("SELECT id, name FROM sites WHERE name=? OR code=?",(site_name,site_name)).fetchone()
            if not site:
                site = db.execute("SELECT id, name FROM sites WHERE name LIKE ?",('%'+site_name+'%',)).fetchone()
            if not site: continue
            sid, sname = site[0], site[1]
            for period, flag, label in [('daily',day_flag,'日巡检方案'),('weekly',week_flag,'周巡检方案'),('monthly',month_flag,'月巡检方案')]:
                if not _yes(flag): continue
                db.execute("INSERT OR IGNORE INTO inspection_schemes (site_id,period,name) VALUES (?,?,?)",(sid,period,f'{sname}-{label}'))
                scheme = db.execute("SELECT id FROM inspection_schemes WHERE site_id=? AND period=?",(sid,period)).fetchone()
                if not scheme: continue
                sc_id = scheme['id']
                # Check if item already exists
                existing = db.execute("SELECT id FROM inspection_scheme_items WHERE scheme_id=? AND check_item=?",(sc_id,item)).fetchone()
                if not existing:
                    next_order = db.execute("SELECT COALESCE(MAX(sort_order),-1)+1 as n FROM inspection_scheme_items WHERE scheme_id=?",(sc_id,)).fetchone()['n']
                    db.execute("INSERT INTO inspection_scheme_items (scheme_id,category,check_item,sort_order) VALUES (?,?,?,?)",(sc_id,cat,item,next_order))
                    created += 1
        db.commit()
    return jsonify({'success':True,'created':created,"warn":"仅记录新增项，已有项未覆盖"})


UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'frontend', 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

@app.route('/api/upload', methods=['POST'])
def upload_file():
    """上传图片/附件，返回可访问的URL。支持 multipart/form-data，字段名 file"""
    file = request.files.get('file')
    if not file: return jsonify({'error':'请选择文件'}),400
    ext = os.path.splitext(file.filename or '.jpg')[1] or '.jpg'
    fname = str(uuid.uuid4())[:8] + ext
    path = os.path.join(UPLOAD_DIR, fname)
    file.save(path)
    # 返回相对路径 URL
    return jsonify({'success':True,'url':'/uploads/'+fname})


@app.route('/api/sites/<int:site_id>/schemes')

@app.route('/api/sites/<int:site_id>/schemes')
def get_site_schemes(site_id):
    with get_db() as db:
        schemes = db.execute("SELECT s.*,(SELECT COUNT(*) FROM inspection_scheme_items i WHERE i.scheme_id=s.id) as item_count FROM inspection_schemes s WHERE s.site_id=? ORDER BY CASE s.period WHEN 'daily' THEN 1 WHEN 'weekly' THEN 2 ELSE 3 END",(site_id,)).fetchall()
        return jsonify([dict(r) for r in schemes])

@app.route('/api/schemes/<int:scheme_id>')
def get_scheme_detail(scheme_id):
    with get_db() as db:
        scheme = db.execute("SELECT * FROM inspection_schemes WHERE id=?",(scheme_id,)).fetchone()
        if not scheme: return jsonify({'error':'方案不存在'}),404
        items = db.execute("SELECT * FROM inspection_scheme_items WHERE scheme_id=? ORDER BY sort_order",(scheme_id,)).fetchall()
        result = dict(scheme); result['items']=[dict(r) for r in items]
        return jsonify(result)

@app.route('/api/schemes/<int:scheme_id>', methods=['PUT'])
def update_scheme(scheme_id):
    data = request.json
    with get_db() as db:
        scheme = db.execute("SELECT * FROM inspection_schemes WHERE id=?",(scheme_id,)).fetchone()
        if not scheme: return jsonify({'error':'方案不存在'}),404
        if 'name' in data: db.execute("UPDATE inspection_schemes SET name=?,updated_at=datetime('now','localtime') WHERE id=?",(data['name'],scheme_id))
        if 'items' in data:
            db.execute("DELETE FROM inspection_scheme_items WHERE scheme_id=?",(scheme_id,))
            for idx,item in enumerate(data['items']):
                db.execute("INSERT INTO inspection_scheme_items (scheme_id,category,check_item,sort_order,is_required) VALUES (?,?,?,?,?)",(scheme_id,item.get('category',''),item.get('check_item',''),idx,item.get('is_required',1)))
        db.commit()
        return jsonify({'success':True})

@app.route('/api/schemes/<int:scheme_id>/items', methods=['POST'])
def add_scheme_item(scheme_id):
    data = request.json
    item_name = data.get('check_item','').strip()
    if not item_name: return jsonify({'error':'检查项不能为空'}),400
    with get_db() as db:
        max_order = db.execute("SELECT COALESCE(MAX(sort_order),-1)+1 as n FROM inspection_scheme_items WHERE scheme_id=?",(scheme_id,)).fetchone()['n']
        db.execute("INSERT INTO inspection_scheme_items (scheme_id,category,check_item,sort_order) VALUES (?,?,?,?)",(scheme_id,data.get('category','自定义'),item_name,max_order))
        db.execute("UPDATE inspection_schemes SET updated_at=datetime('now','localtime') WHERE id=?",(scheme_id,))
        db.commit()
        return jsonify({'success':True})

@app.route('/api/schemes/items/<int:item_id>', methods=['DELETE'])
def delete_scheme_item_ep(item_id):
    with get_db() as db:
        db.execute("DELETE FROM inspection_scheme_items WHERE id=?",(item_id,))
        db.commit()
        return jsonify({'success':True})

@app.route('/api/inspections/auto-generate', methods=['POST'])
def auto_generate_inspections():
    with get_db() as db:
        schemes = db.execute("SELECT * FROM inspection_schemes WHERE status='active'").fetchall()
        generated = 0; today = datetime.now().strftime('%Y-%m-%d')
        for scheme in schemes:
            period = scheme['period']
            if period == 'daily': pass
            elif period == 'weekly' and datetime.now().weekday() != 0: continue
            elif period == 'monthly' and datetime.now().day != 1: continue
            elif period not in ('daily','weekly','monthly'): continue
            existing = db.execute("SELECT id FROM inspection_plans WHERE site_id=? AND scheme_id=? AND start_date=?",(scheme['site_id'],scheme['id'],today)).fetchone()
            if existing: continue
            db.execute("INSERT INTO inspection_plans (plan_name,site_id,type,start_date,end_date,status,scheme_id) VALUES (?,?,?,?,?,?,?)",(scheme['name'],scheme['site_id'],period,today,today,'pending',scheme['id']))
            plan_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            items = db.execute("SELECT * FROM inspection_scheme_items WHERE scheme_id=? ORDER BY sort_order",(scheme['id'],)).fetchall()
            for item in items:
                db.execute("INSERT INTO inspection_tasks (plan_id,site_id,check_item) VALUES (?,?,?)",(plan_id,scheme['site_id'],item['check_item']))
            generated += 1
        db.commit()
        return jsonify({'success':True,'generated':generated})

# --- Workorder management ---@app.route('/api/workorders/<order_no>', methods=['DELETE'])
def delete_workorder(order_no):
    """删除工单（仅支持待受理或已关闭的工单）"""
    with get_db() as db:
        cur = db.execute('SELECT status FROM work_orders WHERE order_no=?', (order_no,)).fetchone()
        if not cur:
            return jsonify({'error': 'not found'}), 404
        if cur['status'] not in ('pending', 'closed'):
            return jsonify({'error': '只能删除待受理或已关闭的工单'}), 400
        db.execute('DELETE FROM work_orders WHERE order_no=?', (order_no,))
        db.execute("DELETE FROM timeline_events WHERE source_type='workorder' AND source_id=?", (order_no,))
        db.commit()
        return jsonify({'success': True})

# --- Maintenance Templates ---
@app.route('/api/maintenance/templates')
def get_maintenance_templates():
    """返回所有运维模板"""
    with get_db() as db:
        rows = db.execute("SELECT * FROM maintenance_templates ORDER BY sort_order").fetchall()
        result = []
        for r in rows:
            d = dict(r)
            if d['check_items']:
                try:
                    d['check_items'] = json.loads(d['check_items'])
                except:
                    d['check_items'] = []
            result.append(d)
        return jsonify(result)

# --- Maintenance Plans ---
@app.route('/api/maintenance/plans')
def get_maintenance_plans():
    with get_db() as db:
        status = request.args.get('status')
        category = request.args.get('category')
        q = "SELECT mp.*, s.name as site_name, s.code as site_code FROM maintenance_plans mp LEFT JOIN sites s ON mp.site_id=s.id"
        params = []
        conds = []
        if status:
            conds.append("mp.status=?")
            params.append(status)
        if category:
            conds.append("mp.category=?")
            params.append(category)
        if conds:
            q += " WHERE " + " AND ".join(conds)
        q += " ORDER BY mp.due_date ASC"
        rows = db.execute(q, params).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/maintenance/plans', methods=['POST'])
def create_maintenance_plan():
    data = request.json
    template_id = data.get('template_id')
    with get_db() as db:
        if template_id:
            # 从模板自动填充
            tpl = db.execute("SELECT * FROM maintenance_templates WHERE id=?", (template_id,)).fetchone()
            if tpl:
                site_name = db.execute("SELECT name FROM sites WHERE id=?", (data['site_id'],)).fetchone()
                site_label = site_name['name'] if site_name else ''
                plan_name = data.get('plan_name') or f"{tpl['title']}-{site_label}"
                category = tpl['category']
                frequency = tpl['frequency']
                sub_category = tpl['sub_category']
                remark = data.get('remark') or tpl['description']
                cur = db.execute(
                    "INSERT INTO maintenance_plans (site_id,plan_name,category,frequency,due_date,assignee,template_id,sub_category,remark) VALUES (?,?,?,?,?,?,?,?,?)",
                    (data['site_id'], plan_name, category, frequency, data.get('due_date'), data.get('assignee'), template_id, sub_category, remark)
                )
            else:
                return jsonify({'error': 'template not found'}), 404
        else:
            # 无模板的传统创建方式
            cur = db.execute(
                "INSERT INTO maintenance_plans (site_id,plan_name,category,frequency,due_date,assignee) VALUES (?,?,?,?,?,?)",
                (data['site_id'], data['plan_name'], data['category'], data.get('frequency','monthly'), data.get('due_date'), data.get('assignee'))
            )
        db.commit()
        return jsonify({'success': True, 'plan_id': cur.lastrowid})

@app.route('/api/maintenance/plans/<int:plan_id>/complete', methods=['PUT'])
def complete_maintenance_plan(plan_id):
    data = request.json or {}
    check_results = data.get('check_results')
    with get_db() as db:
        if check_results:
            db.execute("UPDATE maintenance_plans SET status='completed', completed_at=datetime('now','localtime'), check_results=? WHERE id=?",
                       (json.dumps(check_results, ensure_ascii=False), plan_id))
        else:
            db.execute("UPDATE maintenance_plans SET status='completed', completed_at=datetime('now','localtime') WHERE id=?", (plan_id,))
        db.commit()
        # 记录时间线
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator) VALUES (?,?,?,?)",
                   ('maintenance', plan_id, 'completed', '系统'))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/maintenance/plans/<int:plan_id>/urge', methods=['POST'])
def urge_maintenance(plan_id):
    with get_db() as db:
        db.execute("UPDATE maintenance_plans SET urge_count=COALESCE(urge_count,0)+1, last_urged_at=datetime('now','localtime') WHERE id=?", (plan_id,))
        db.commit()
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator) VALUES (?,?,?,?)",
                   ('maintenance', plan_id, 'urged', '系统'))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/maintenance/stats')
def maintenance_stats():
    """运维统计：今日待办总数/各分类统计"""
    with get_db() as db:
        today = datetime.now().strftime('%Y-%m-%d')
        total_pending = db.execute("SELECT COUNT(*) as c FROM maintenance_plans WHERE status='pending'").fetchone()['c']
        overdue = db.execute("SELECT COUNT(*) as c FROM maintenance_plans WHERE status='pending' AND due_date < ?", (today,)).fetchone()['c']
        return jsonify({'total_pending': total_pending, 'overdue': overdue})

@app.route('/api/maintenance/plans/<int:plan_id>', methods=['PUT'])
def update_maintenance_plan(plan_id):
    """修改运维计划"""
    data = request.json
    with get_db() as db:
        cur = db.execute("SELECT * FROM maintenance_plans WHERE id=?", (plan_id,)).fetchone()
        if not cur:
            return jsonify({'error': 'not found'}), 404
        updates = []
        params = []
        for field in ['plan_name','category','frequency','due_date','site_id','assignee']:
            if field in data:
                updates.append(f"{field}=?")
                params.append(data[field])
        if not updates:
            return jsonify({'error': 'no fields to update'}), 400
        params.append(plan_id)
        db.execute(f"UPDATE maintenance_plans SET {','.join(updates)} WHERE id=?", params)
        db.commit()
        return jsonify({'success': True})

@app.route('/api/maintenance/plans/<int:plan_id>', methods=['DELETE'])
def delete_maintenance_plan(plan_id):
    """删除运维计划"""
    with get_db() as db:
        db.execute("DELETE FROM maintenance_plans WHERE id=?", (plan_id,))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/maintenance/plans/<int:plan_id>/review', methods=['PUT'])
def review_maintenance_plan(plan_id):
    """审核运维计划"""
    data = request.json
    review_result = data.get('review_result', 'pending')
    review_comment = data.get('review_comment', '')
    operator = data.get('operator', '系统')
    with get_db() as db:
        db.execute("UPDATE maintenance_plans SET review_status=?, review_comment=? WHERE id=?",
                   (review_result, review_comment, plan_id))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('maintenance', plan_id, 'reviewed', operator, f'审核结果:{review_result} 意见:{review_comment}'))
        db.commit()
        return jsonify({'success': True})

# --- Water Level Checks (Phase A-3) ---
@app.route('/api/water-level/checks')
def get_water_level_checks():
    """水位差值校验列表"""
    site_id = request.args.get('site_id', '', type=int)
    limit = request.args.get('limit', 50, type=int)
    with get_db() as db:
        q = "SELECT w.*, s.name as site_name, s.code as site_code FROM water_level_checks w LEFT JOIN sites s ON w.site_id=s.id WHERE 1=1"
        params = []
        if site_id:
            q += " AND w.site_id=?"
            params.append(site_id)
        q += " ORDER BY w.created_at DESC LIMIT ?"
        params.append(limit)
        return jsonify([dict(r) for r in db.execute(q, params).fetchall()])

@app.route('/api/water-level/checks', methods=['POST'])
def create_water_level_check():
    """手动录入水位校验"""
    data = request.json
    site_id = data.get('site_id')
    manual_level = data.get('manual_level')
    telemetry_level = data.get('telemetry_level')
    operator = data.get('operator', '系统')
    diff = round(abs(manual_level - telemetry_level), 3)
    status = 'abnormal' if diff > 0.02 else 'normal'
    adjust_action = data.get('adjust_action', '')
    with get_db() as db:
        db.execute("""
            INSERT INTO water_level_checks (site_id,manual_level,telemetry_level,diff,status,adjust_action,operator)
            VALUES (?,?,?,?,?,?,?)
        """, (site_id, manual_level, telemetry_level, diff, status, adjust_action, operator))
        wlc_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        # 如果差值超标，生成告警
        if diff > 0.02:
            site = db.execute("SELECT name, code FROM sites WHERE id=?", (site_id,)).fetchone()
            site_name = site['name'] if site else f'站点{site_id}'
            level = 'red' if diff > 0.05 else 'orange'
            db.execute("""
                INSERT INTO alerts (site_id,level,metric,value,message,status)
                VALUES (?,?,?,?,?,?)
            """, (site_id, level, 'water_level_diff', diff,
                  f'{site_name}水位校验差值{diff}m超过阈值', 'pending'))
            # 时间线
            db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                       ('water_level', wlc_id, 'alert_generated', operator, f'水位差值{diff}m触发告警'))
        db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                   ('water_level', wlc_id, 'checked', operator, f'录入校验-人工{manual_level}m/遥测{telemetry_level}m/差值{diff}m'))
        db.commit()
        return jsonify({'success': True, 'id': wlc_id, 'diff': diff, 'status': status})

@app.route('/api/water-level/checks/auto', methods=['POST'])
def auto_water_level_check():
    """自动生成水位校验记录（模拟手工录入与遥测数据的比较）"""
    with get_db() as db:
        # 选取水位站/水文站
        sites = db.execute(
            "SELECT id, name FROM sites WHERE type IN ('water_level','hydrology') ORDER BY RANDOM() LIMIT 5"
        ).fetchall()
        results = []
        for site in sites:
            manual = round(random.uniform(17.0, 24.0), 2)
            telemetry = round(manual + random.uniform(-0.03, 0.03), 2)
            diff = round(abs(manual - telemetry), 3)
            status = 'abnormal' if diff > 0.02 else 'normal'
            db.execute("""
                INSERT INTO water_level_checks (site_id,manual_level,telemetry_level,diff,status,operator)
                VALUES (?,?,?,?,?,?)
            """, (site['id'], manual, telemetry, diff, status, '自动'))
            wlc_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            if diff > 0.02:
                level = 'red' if diff > 0.05 else 'orange'
                db.execute("""
                    INSERT INTO alerts (site_id,level,metric,value,message,status)
                    VALUES (?,?,?,?,?,?)
                """, (site['id'], level, 'water_level_diff', diff,
                      f'{site["name"]}自动校验差值{diff}m超过阈值', 'pending'))
                db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                           ('water_level', wlc_id, 'alert_generated', '自动', f'自动校验差值{diff}m触发告警'))
            db.execute("INSERT INTO timeline_events (source_type,source_id,event_type,operator,remark) VALUES (?,?,?,?,?)",
                       ('water_level', wlc_id, 'auto_checked', '自动', f'自动校验-人工{manual}m/遥测{telemetry}m/差值{diff}m'))
            results.append({'site_id': site['id'], 'site_name': site['name'], 'diff': diff, 'status': status})
        db.commit()
        return jsonify({'success': True, 'count': len(results), 'results': results})

# --- Data Arrival ---
@app.route('/api/data/arrival')
def get_data_arrival():
    """当日到报率数据"""
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    with get_db() as db:
        rows = db.execute(
            "SELECT da.*, s.name as site_name, s.code as site_code, s.type as site_type FROM data_arrival da LEFT JOIN sites s ON da.site_id=s.id WHERE da.date=? ORDER BY da.arrival_rate ASC",
            (date,)
        ).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/data/arrival/summary')
def data_arrival_summary():
    """到报率汇总：按项目分类"""
    date = request.args.get('date', datetime.now().strftime('%Y-%m-%d'))
    with get_db() as db:
        # 检查是否有今天的数据，没有则尝试从sensor_data实时计算
        has_data = db.execute("SELECT COUNT(*) as c FROM data_arrival WHERE date=?", (date,)).fetchone()['c']
        if has_data == 0:
            # 回退：从sensor_data统计今天的到报情况
            rows = db.execute("""
                SELECT s.type as metric, 
                    COUNT(DISTINCT sd.site_id) as site_count,
                    ROUND(AVG(CASE WHEN sd.id IS NOT NULL THEN 100.0 ELSE 0 END),1) as avg_rate,
                    0 as below_threshold
                FROM sites s
                LEFT JOIN sensor_data sd ON s.id = sd.site_id AND sd.recorded_at >= ?
                GROUP BY s.type
            """, (date + ' 00:00:00',)).fetchall()
            # 如果回退也无数据，返回空
            if not rows or all((r['avg_rate'] or 0) == 0 for r in rows):
                return jsonify({'total_avg': 0, 'by_metric': []})
            total_avg = round(sum(r['avg_rate'] or 0 for r in rows) / max(len(rows), 1), 1)
            return jsonify({
                'total_avg': total_avg,
                'by_metric': [dict(r) for r in rows]
            })
        total = db.execute("SELECT AVG(arrival_rate) as avg FROM data_arrival WHERE date=?", (date,)).fetchone()
        return jsonify({
            'total_avg': round(total['avg'],1) if total and total['avg'] else 0,
            'by_metric': [dict(r) for r in rows] if has_data > 0 else []
        })

# --- Hotline ---
@app.route('/api/hotline/events')
def get_hotline_events():
    limit = request.args.get('limit', 50, type=int)
    with get_db() as db:
        rows = db.execute("SELECT * FROM hotline_events ORDER BY created_at DESC LIMIT ?", (limit,)).fetchall()
        return jsonify([dict(r) for r in rows])

@app.route('/api/hotline/events', methods=['POST'])
def create_hotline_event():
    data = request.json
    with get_db() as db:
        db.execute("""
            INSERT INTO hotline_events (caller_name,caller_phone,event_type,description,location,operator)
            VALUES (?,?,?,?,?,?)
        """, (data.get('caller_name',''),data.get('caller_phone',''),
              data.get('event_type',''),data.get('description',''),
              data.get('location',''),data.get('operator','')))
        db.commit()
        return jsonify({'success': True})

@app.route('/api/hotline/events/<int:event_id>/convert', methods=['POST'])
def convert_hotline_to_order(event_id):
    """热线事件转工单"""
    data = request.json
    with get_db() as db:
        event = db.execute("SELECT * FROM hotline_events WHERE id=?", (event_id,)).fetchone()
        if not event:
            return jsonify({'error': 'not found'}), 404
        now = datetime.now()
        order_no = f"WO-{now.strftime('%Y%m%d')}-{random.randint(100,999)}"
        leve = data.get('level','normal')
        sla_hours = {'normal': 72, 'urgent': 24, 'critical': 2}.get(leve, 72)
        sla_deadline = (now + timedelta(hours=sla_hours)).strftime('%Y-%m-%d %H:%M')
        db.execute("""
            INSERT INTO work_orders (order_no,source,event_type,level,title,description,assignee,status,sla_deadline)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (order_no, 'hotline', event['event_type'], leve,
              f"[热线]{event['event_type']}", event['description'],
              data.get('assignee',''), 'pending', sla_deadline))
        db.execute("UPDATE hotline_events SET status='dispatched', related_order_no=? WHERE id=?", (order_no, event_id))
        db.commit()
        return jsonify({'success': True, 'order_no': order_no})

# --- Dashboard ---
@app.route('/api/dashboard/summary')
@login_required
def dashboard_summary():
    with get_db() as db:
        overview = {
            'total_sites': db.execute("SELECT COUNT(*) as c FROM sites").fetchone()['c'],
            'online_sites': db.execute("SELECT COUNT(*) as c FROM sites WHERE status='online'").fetchone()['c'],
            'device_total': db.execute("SELECT COUNT(*) as c FROM device_shadows").fetchone()['c'],
            'device_online': db.execute("SELECT COUNT(*) as c FROM device_shadows WHERE status='online'").fetchone()['c'],
            'active_alerts': db.execute("SELECT COUNT(*) as c FROM alerts WHERE status='pending'").fetchone()['c'],
            'open_orders': db.execute("SELECT COUNT(*) as c FROM work_orders WHERE status NOT IN ('closed')").fetchone()['c'],
            'today_orders': db.execute("SELECT COUNT(*) as c FROM work_orders WHERE date(created_at)=date('now','localtime')").fetchone()['c'],
        }
        # 最新告警
        latest_alerts = db.execute("""
            SELECT a.*, s.name as site_name FROM alerts a LEFT JOIN sites s ON a.site_id=s.id
            WHERE a.status='pending' ORDER BY CASE level WHEN 'red' THEN 1 WHEN 'orange' THEN 2 ELSE 3 END, a.created_at DESC LIMIT 5
        """).fetchall()
        # 待处理工单
        pending_orders = db.execute("""
            SELECT w.*, s.name as site_name FROM work_orders w LEFT JOIN sites s ON w.site_id=s.id
            WHERE w.status NOT IN ('closed') ORDER BY w.created_at DESC LIMIT 5
        """).fetchall()
        # 今日巡检
        insp = db.execute("""
            SELECT COUNT(*) as today_total,
                (SELECT COUNT(*) FROM inspection_tasks WHERE date(check_time)=date('now','localtime') AND result='abnormal') as abnormal
            FROM inspection_tasks WHERE date(check_time)=date('now','localtime')
        """).fetchone()
        return jsonify({
            'overview': overview,
            'latest_alerts': [dict(a) for a in latest_alerts],
            'pending_orders': [dict(o) for o in pending_orders],
            'today_inspection_total': insp['today_total'] or 0,
            'today_inspection_abnormal': insp['abnormal'] or 0,
        })

# --- 实时天气获取 ---
# 默认坐标：南昌（可修改为项目所在地）
WEATHER_LAT = 28.68
WEATHER_LON = 115.89

def fetch_real_weather():
    """从 Open-Meteo 免费接口获取实时天气，写入数据库"""
    url = (
        f"https://api.open-meteo.com/v1/forecast?"
        f"latitude={WEATHER_LAT}&longitude={WEATHER_LON}"
        f"&current=temperature_2m,relative_humidity_2m,precipitation,pressure_msl,"
        f"wind_speed_10m,wind_direction_10m,weather_code"
        f"&timezone=auto"
    )
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'WaterOps/1.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = _json.loads(resp.read().decode())
    except Exception as e:
        print(f"[Weather] 获取天气失败: {e}")
        return False

    current = data.get('current', {})
    if not current:
        return False

    # WMO 天气代码 → 中文天气类型
    wmo_map = {
        0:'晴',1:'晴',2:'多云',3:'阴',45:'雾',48:'雾',
        51:'小雨',53:'小雨',55:'中雨',56:'冻雨',57:'冻雨',
        61:'小雨',63:'中雨',65:'大雨',66:'冻雨',67:'冻雨',
        71:'小雪',73:'中雪',75:'大雪',77:'雪粒',
        80:'阵雨',81:'中雨',82:'大雨',85:'阵雪',86:'阵雪',
        95:'雷雨',96:'雷雨',99:'雷雨'
    }
    wcode = current.get('weather_code', 0)
    weather_type = wmo_map.get(wcode, '多云')

    # 温度(°C), 湿度(%), 降水量(mm), 气压(hPa), 风速(km/h)
    temp = current.get('temperature_2m', 25)
    humidity = current.get('relative_humidity_2m', 60)
    precip = current.get('precipitation', 0)
    pressure = current.get('pressure_msl', 1013)
    wind_speed = current.get('wind_speed_10m', 5)
    wind_dir_deg = current.get('wind_direction_10m', 0)

    # 风向角度 → 中文
    dirs = ['北','东北','东','东南','南','西南','西','西北']
    wind_dir = dirs[round(wind_dir_deg / 45) % 8]

    # 生成预警信息
    warnings = []
    if precip > 10: warnings.append('暴雨')
    if wind_speed > 40: warnings.append('大风')
    if temp > 35: warnings.append('高温')

    with get_db() as db:
        # 清理旧数据，只保留最新一条
        db.execute("DELETE FROM weather_data WHERE id NOT IN (SELECT id FROM weather_data ORDER BY id DESC LIMIT 1)")
        db.execute(
            """INSERT INTO weather_data (temperature,humidity,wind_speed,wind_direction,
               precipitation,pressure,weather_type,warning_info,recorded_at)
               VALUES (?,?,?,?,?,?,?,?,datetime('now','localtime'))""",
            (temp, humidity, wind_speed, wind_dir, precip, pressure, weather_type,
             ','.join(warnings) if warnings else '')
        )
        db.commit()
    print(f"[Weather] 已更新: {weather_type} {temp}°C 降水{precip}mm 风速{wind_speed}km/h")
    return True

# --- 天气数据 ---
@app.route('/api/weather')
def get_weather():
    """返回当前天气数据、未来24小时逐时预报、天气预警"""
    with get_db() as db:
        # 获取最新天气记录
        current = db.execute(
            "SELECT * FROM weather_data ORDER BY recorded_at DESC LIMIT 1"
        ).fetchone()

        # 无数据或数据超过5分钟 → 刷新（确保移动端/PC端看到一致的最新数据）
        need_fetch = False
        if not current:
            need_fetch = True
        else:
            try:
                from datetime import datetime as _dt
                last = _dt.strptime(current['recorded_at'], '%Y-%m-%d %H:%M:%S')
                if (_dt.now() - last).total_seconds() > 300:  # 5分钟
                    need_fetch = True
            except:
                need_fetch = True

        if need_fetch:
            fetch_real_weather()
            # 重新读取
            current = db.execute(
                "SELECT * FROM weather_data ORDER BY recorded_at DESC LIMIT 1"
            ).fetchone()

        if not current:
            return jsonify({'error': '暂无天气数据，请稍后重试'}), 503

        # 构建当前天气
        now = datetime.now()
        result = {
            'current': {
                'temperature': current['temperature'],
                'humidity': current['humidity'],
                'wind_speed': current['wind_speed'],
                'wind_direction': current['wind_direction'],
                'precipitation': current['precipitation'],
                'pressure': current['pressure'],
                'weather_type': current['weather_type'],
                'recorded_at': current['recorded_at'],
            },
            'hourly_forecast': [],  # 未来24小时逐小时预报
            'warnings': [],  # 天气预警
        }

        # 解析预警信息
        if current['warning_info']:
            for w in current['warning_info'].split(','):
                if '暴雨' in w:
                    result['warnings'].append({'type': '暴雨', 'level': '黄色', 'message': '预计未来6小时有暴雨，请加强防范'})
                elif '大风' in w:
                    result['warnings'].append({'type': '大风', 'level': '蓝色', 'message': '风力已达6级以上，请加固设施'})
                elif '高温' in w:
                    result['warnings'].append({'type': '高温', 'level': '橙色', 'message': '最高气温超过35℃，请做好防暑降温'})

        # 生成未来24小时逐小时预报
        directions = ['北', '东北', '东', '东南', '南', '西南', '西', '西北']
        weather_types = ['晴', '多云', '阴', '小雨', '中雨', '大雨']
        for i in range(24):
            forecast_time = now + timedelta(hours=i+1)
            # 基于当前值做小幅随机波动
            temp_forecast = round(current['temperature'] + random.uniform(-3, 3), 1)
            hum_forecast = round(current['humidity'] + random.uniform(-10, 10), 1)
            hum_forecast = max(20, min(100, hum_forecast))
            wind_forecast = round(current['wind_speed'] + random.uniform(-2, 2), 1)
            wind_forecast = max(0, wind_forecast)
            precip_forecast = round(current['precipitation'] * random.uniform(0.5, 1.5), 1)

            result['hourly_forecast'].append({
                'time': forecast_time.strftime('%H:%M'),
                'datetime': forecast_time.strftime('%Y-%m-%d %H:%M'),
                'temperature': temp_forecast,
                'humidity': hum_forecast,
                'wind_speed': wind_forecast,
                'wind_direction': random.choice(directions),
                'precipitation': precip_forecast,
                'weather': random.choices(weather_types,
                    weights=[0.35, 0.25, 0.15, 0.1, 0.1, 0.05])[0],
            })

        return jsonify(result)

# --- 降雨预报 ---
@app.route('/api/rainfall/forecast')
def rainfall_forecast():
    """降雨预报数据：当前降雨 + 未来48小时逐小时预报 + 数据来源"""
    with get_db() as db:
        # 当前实时降雨（取最新天气记录的降水量）
        current = db.execute("SELECT precipitation, weather_type, temperature FROM weather_data ORDER BY id DESC LIMIT 1").fetchone()
        now = datetime.now()
        # 模拟48小时逐小时降雨预报
        hours = []
        base_precip = current['precipitation'] if current else 5.0
        base_weather = current['weather_type'] if current else '多云'
        rainy_weights = [0.6, 0.2, 0.1, 0.05, 0.05]
        rain_types = ['无雨', '小雨', '中雨', '大雨', '暴雨']
        for i in range(48):
            t = now + timedelta(hours=i)
            # 不同时段降雨概率不同
            hour_of_day = t.hour
            if 2 <= hour_of_day <= 6:
                prob = 0.15  # 凌晨降雨概率低
            elif 14 <= hour_of_day <= 17:
                prob = 0.40  # 午后对流高
            else:
                prob = 0.25
            is_rain = random.random() < prob
            if is_rain:
                p = round(random.uniform(0.5, max(2, base_precip * 1.5)), 1)
                wt = random.choices(rain_types[1:], weights=[0.5, 0.3, 0.15, 0.05])[0]
            else:
                p = 0
                wt = '无雨'
            hours.append({
                'time': t.strftime('%m-%d %H:00'),
                'precipitation': p,
                'rain_type': wt,
                'probability': round(prob * 100),
            })
        sources = ['自动监测站', '气象局', '雷达估测']
        return jsonify({
            'current_rainfall': round(base_precip, 1) if base_precip else 0,
            'current_weather': base_weather,
            'forecast': hours,
            'sources': sources,
        })

# --- 水质监测 ---
@app.route('/api/water-quality')
def water_quality():
    """水质监测数据：返回各供水站/水库的水质指标及7日均值对比"""
    site_id = request.args.get('site_id', type=int)
    with get_db() as db:
        # 查询所有水源相关站点（水库 + 供水站）
        q = "SELECT * FROM sites WHERE type IN ('reservoir','water_supply')"
        params = []
        if site_id:
            q += " AND id=?"
            params.append(site_id)
        sites = db.execute(q, params).fetchall()

        if not sites:
            return jsonify({'error': '没有找到水源相关站点'}), 404

        # 水质指标定义
        water_metrics = [
            ('turbidity', 'NTU', '浊度'),
            ('ph', '', 'pH值'),
            ('chlorine', 'mg/L', '余氯'),
            ('ammonia', 'mg/L', '氨氮'),
            ('cod', 'mg/L', 'COD'),
        ]
        seven_days_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')

        result = []
        for site in sites:
            site_data = {
                'site_id': site['id'],
                'site_code': site['code'],
                'site_name': site['name'],
                'site_type': site['type'],
                'metrics': {},
            }
            for metric, unit, label in water_metrics:
                # 最新值 + 阈值
                latest = db.execute(
                    "SELECT value, threshold_high, threshold_critical, recorded_at FROM sensor_data WHERE site_id=? AND metric=? ORDER BY recorded_at DESC LIMIT 1",
                    (site['id'], metric)
                ).fetchone()

                # 7天均值
                avg_row = db.execute(
                    "SELECT AVG(value) as avg_val FROM sensor_data WHERE site_id=? AND metric=? AND recorded_at > ?",
                    (site['id'], metric, seven_days_ago)
                ).fetchone()

                site_data['metrics'][metric] = {
                    'label': label,
                    'unit': unit,
                    'current': round(latest['value'], 3) if latest and latest['value'] is not None else None,
                    'current_time': latest['recorded_at'] if latest else None,
                    'avg_7d': round(avg_row['avg_val'], 3) if avg_row and avg_row['avg_val'] is not None else None,
                    'thresh_high': round(latest['threshold_high'], 3) if latest and latest['threshold_high'] is not None else None,
                    'thresh_critical': round(latest['threshold_critical'], 3) if latest and latest['threshold_critical'] is not None else None,
                }

            result.append(site_data)

        return jsonify(result)

# --- 设备状态监控 (新增) ---
@app.route('/api/devices/status')
@login_required
def device_status():
    """设备心跳状态汇总：在线/离线统计、各类型统计、离线设备明细"""
    with get_db() as db:
        # 设备总数与状态统计
        total = db.execute("SELECT COUNT(*) as c FROM device_shadows").fetchone()['c']
        online = db.execute("SELECT COUNT(*) as c FROM device_shadows WHERE status='online'").fetchone()['c']
        offline = db.execute("SELECT COUNT(*) as c FROM device_shadows WHERE status='offline'").fetchone()['c']

        # 按设备类型统计
        by_type = db.execute("""
            SELECT device_type,
                COUNT(*) as total,
                SUM(CASE WHEN status='online' THEN 1 ELSE 0 END) as online,
                SUM(CASE WHEN status='offline' THEN 1 ELSE 0 END) as offline
            FROM device_shadows GROUP BY device_type ORDER BY device_type
        """).fetchall()

        # 离线设备明细
        offline_devices = db.execute("""
            SELECT d.*, s.name as site_name, s.code as site_code
            FROM device_shadows d LEFT JOIN sites s ON d.site_id=s.id
            WHERE d.status='offline'
            ORDER BY d.last_data_time DESC
        """).fetchall()

        # 各站点设备状态
        site_devices = db.execute("""
            SELECT s.id as site_id, s.code, s.name,
                COUNT(d.id) as total_devices,
                SUM(CASE WHEN d.status='online' THEN 1 ELSE 0 END) as online_devices,
                SUM(CASE WHEN d.status='offline' THEN 1 ELSE 0 END) as offline_devices
            FROM sites s LEFT JOIN device_shadows d ON s.id=d.site_id
            GROUP BY s.id ORDER BY s.id
        """).fetchall()

        return jsonify({
            'summary': {
                'total': total,
                'online': online,
                'offline': offline,
                'online_rate': round(online / total * 100, 1) if total > 0 else 0,
            },
            'by_type': [dict(r) for r in by_type],
            'offline_devices': [dict(d) for d in offline_devices],
            'site_devices': [dict(s) for s in site_devices],
        })

# --- 数据质量报告 (新增) ---
@app.route('/api/data-quality')
def data_quality():
    """数据质量报告：今日到达率/完整率/及时率、异常站点、24小时趋势"""
    with get_db() as db:
        today = datetime.now().strftime('%Y-%m-%d')
        now = datetime.now()
        twenty_four_hours_ago = (now - timedelta(hours=24)).strftime('%Y-%m-%d %H:%M:%S')

        # 获取所有站点及其类型，确定每个站点期望的指标列表
        sites = db.execute("SELECT * FROM sites").fetchall()
        expected_metrics = {
            'rainfall': ['rainfall', 'precipitation', 'cumulative_rainfall'],
            'water_level': ['water_level', 'flow', 'velocity'],
            'hydrology': ['water_level', 'rainfall', 'flow', 'velocity', 'sediment'],
            'soil_moisture': ['soil_moisture', 'soil_temperature'],
            'evaporation': ['evaporation', 'temperature', 'wind_speed'],
        }

        # 计算今日数据到达率和完整率
        total_expected = 0
        total_received = 0
        anomaly_sites = []  # 异常站点列表

        for site in sites:
            site_metrics = expected_metrics.get(site['type'], [])
            if not site_metrics:
                continue

            # 期望数据点：按每小时12条(每5分钟一条)估算，匹配后端的回填频率
            hours_elapsed = max(1, (now.hour * 60 + now.minute) / 60)
            expected_per_metric = max(1, int(hours_elapsed * 12))  # 每小时12条(每5分钟一次)
            expected_today = len(site_metrics) * expected_per_metric
            total_expected += expected_today

            # 实际收到数据条数
            received = db.execute(
                "SELECT COUNT(*) as c FROM sensor_data WHERE site_id=? AND recorded_at LIKE ?",
                (site['id'], today + '%')
            ).fetchone()['c']
            total_received += received

            # 计算该站点数据完整率
            completeness = round(received / expected_today * 100, 1) if expected_today > 0 else 0

            # 标注异常站点（数据到达率<50%或数据为0的站点）
            if completeness < 50 and expected_today > 5:
                anomaly_sites.append({
                    'site_id': site['id'],
                    'site_code': site['code'],
                    'site_name': site['name'],
                    'site_type': site['type'],
                    'expected': expected_today,
                    'received': received,
                    'completeness': completeness,
                    'reason': '数据到达率低于50%' if received < expected_today / 2 else '数据采集异常',
                })

        # 总体指标
        arrival_rate = round(min(100, total_received / total_expected * 100), 1) if total_expected > 0 else 0
        completeness_rate = arrival_rate  # 在模拟场景下，到达率≈完整率

        # 数据及时率 (最近1小时内是否有数据视为及时)
        one_hour_ago = (now - timedelta(hours=1)).strftime('%Y-%m-%d %H:%M:%S')
        timely_sites = db.execute("""
            SELECT COUNT(DISTINCT site_id) as c FROM sensor_data WHERE recorded_at > ?
        """, (one_hour_ago,)).fetchone()['c']
        total_sites = len(sites) if sites else 1
        timeliness_rate = round(timely_sites / total_sites * 100, 1)

        # 最近24小时数据质量趋势 (按小时分组)
        hourly_trend = db.execute("""
            SELECT strftime('%Y-%m-%d %H:00', recorded_at) as hour,
                COUNT(*) as data_count
            FROM sensor_data
            WHERE recorded_at > ?
            GROUP BY strftime('%Y-%m-%d %H:00', recorded_at)
            ORDER BY hour ASC
        """, (twenty_four_hours_ago,)).fetchall()

        # 每个小时期望数据量（所有站点 × 12条/小时×指标数）
        expected_per_hour = sum(len(expected_metrics.get(s['type'], [])) for s in sites) * 12

        trend = []
        for row in hourly_trend:
            trend.append({
                'hour': row['hour'],
                'count': row['data_count'],
                'rate': round(min(100, row['data_count'] / expected_per_hour * 100), 1) if expected_per_hour > 0 else 0,
            })

        return jsonify({
            'today': {
                'arrival_rate': arrival_rate,         # 数据到达率(%)
                'completeness_rate': completeness_rate, # 数据完整率(%)
                'timeliness_rate': timeliness_rate,     # 数据及时率(%)
                'total_expected': total_expected,
                'total_received': total_received,
                'active_sites': total_sites,
                'timely_sites': timely_sites,
            },
            'anomaly_sites': anomaly_sites,
            'hourly_trend': trend,
        })

# ===================== 数据回填（生成历史数据用于图表展示） =====================

def backfill_history(hours=72):
    """回填历史监测数据，让图表有历史数据可展示"""
    with get_db() as db:
        count = db.execute("SELECT COUNT(*) as c FROM sensor_data").fetchone()['c']
        if count > 10000:
            print(f"[Backfill] 已有 {count} 条数据，跳过回填")
            return
        sites = db.execute("SELECT * FROM sites").fetchall()
        now = datetime.now()
        print(f"[Backfill] 开始回填 {hours} 小时历史数据...")
        for h in range(hours * 12, 0, -1):  # 每5分钟一条，共hours小时
            ts = (now - timedelta(minutes=5 * h)).strftime('%Y-%m-%d %H:%M:%S')
            for site in sites:
                sid = site['id']
                stype = site['type']
                if stype == 'reservoir':
                    wl = round(random.uniform(46.0, 52.5), 2)
                    sp = round(random.uniform(0.05, 0.6), 3)
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'water_level',wl,'m',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'seepage',sp,'L/s',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'inflow',round(random.uniform(10,50),1),'m³/s',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'outflow',round(random.uniform(5,40),1),'m³/s',ts))
                elif stype == 'sluice':
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'gate_opening',round(random.uniform(20,80),1),'%',ts))
                elif stype == 'dike':
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'displacement',round(random.uniform(0,12),1),'mm',ts))
                elif stype == 'pump':
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'flow',round(random.uniform(0.5,3.5),1),'m³/s',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'vibration',round(random.uniform(2,9),1),'mm/s',ts))
                elif stype == 'water_supply':
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'turbidity',round(random.uniform(0.05,0.6),2),'NTU',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'ph',round(random.uniform(6.8,7.8),1),'',ts))
                elif stype == 'irrigation':
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'water_level',round(random.uniform(2.0,4.5),1),'m',ts))
                    db.execute("INSERT INTO sensor_data (site_id,metric,value,unit,recorded_at) VALUES (?,?,?,?,?)",(sid,'soil_moisture',round(random.uniform(30,90),1),'%',ts))
        db.commit()
        total = db.execute("SELECT COUNT(*) as c FROM sensor_data").fetchone()['c']
        print(f"[Backfill] 完成！共 {total} 条历史数据")


# ===================== 认证API端点 =====================

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data = request.get_json() or {}
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    if not username or not password:
        return jsonify({'error': '请输入用户名和密码'}), 400
    with get_db() as db:
        user = db.execute("SELECT id, username, password_hash, role, real_name, phone, status FROM users WHERE username=? AND status='active'",
                          (username,)).fetchone()
    if not user or user['password_hash'] != _hash_pw(password):
        return jsonify({'error': '用户名或密码错误'}), 401
    token = secrets.token_urlsafe(32)
    _tokens[token] = {
        'id': user['id'],
        'username': user['username'],
        'role': user['role'],
        'real_name': user['real_name'],
        'phone': user['phone'] or '',
    }
    # 获取此用户可管理的站点列表
    with get_db() as db:
        site_rows = db.execute("SELECT s.id, s.name, s.code, s.type FROM sites s JOIN user_sites us ON s.id=us.site_id WHERE us.user_id=?", (user['id'],)).fetchall()
    sites = [{'id': r['id'], 'name': r['name'], 'code': r['code'], 'type': r['type']} for r in site_rows]
    return jsonify({
        'success': True,
        'token': token,
        'user': {
            'id': user['id'],
            'username': user['username'],
            'role': user['role'],
            'real_name': user['real_name'],
            'phone': user['phone'] or '',
        },
        'sites_count': len(sites),
        'sites': sites,
    })

@app.route('/api/auth/me')
@login_required
def api_me():
    return jsonify({
        'success': True,
        'user': g.current_user,
        'site_ids': g.user_site_ids,
    })


# ===================== 用户管理API（管理员） =====================

@app.route('/api/users')
@login_required
def api_users():
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '无权限'}), 403
    with get_db() as db:
        rows = db.execute("SELECT id, username, role, real_name, phone, status, created_at FROM users ORDER BY id").fetchall()
    users = []
    for r in rows:
        with get_db() as db2:
            cnt = db2.execute("SELECT COUNT(*) as c FROM user_sites WHERE user_id=?", (r['id'],)).fetchone()['c']
        users.append({
            'id': r['id'], 'username': r['username'], 'role': r['role'],
            'real_name': r['real_name'], 'phone': r['phone'] or '',
            'status': r['status'], 'sites_count': cnt,
            'created_at': r['created_at'],
        })
    return jsonify(users)

@app.route('/api/users/<int:uid>/sites', methods=['GET'])
@login_required
def api_user_sites(uid):
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '无权限'}), 403
    with get_db() as db:
        sids = [r['site_id'] for r in db.execute("SELECT site_id FROM user_sites WHERE user_id=?", (uid,)).fetchall()]
    return jsonify({'site_ids': sids})

@app.route('/api/users/<int:uid>/sites', methods=['PUT'])
@login_required
def api_update_user_sites(uid):
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '无权限'}), 403
    data = request.get_json() or {}
    site_ids = data.get('site_ids', [])
    if not isinstance(site_ids, list):
        return jsonify({'error': 'site_ids格式错误'}), 400
    with get_db() as db:
        db.execute("DELETE FROM user_sites WHERE user_id=?", (uid,))
        for sid in site_ids:
            db.execute("INSERT OR IGNORE INTO user_sites (user_id,site_id) VALUES (?,?)", (uid, sid))
        db.commit()
    return jsonify({'success': True, 'count': len(site_ids)})

@app.route('/api/users/<int:uid>/reset-password', methods=['PUT'])
@login_required
def api_reset_password(uid):
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '无权限'}), 403
    data = request.get_json() or {}
    new_pw = data.get('new_password', 'yw123456')
    with get_db() as db:
        db.execute("UPDATE users SET password_hash=? WHERE id=?", (_hash_pw(new_pw), uid))
        db.commit()
    return jsonify({'success': True})

@app.route('/api/users/<int:uid>/status', methods=['PUT'])
@login_required
def api_user_status(uid):
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '无权限'}), 403
    data = request.get_json() or {}
    new_status = data.get('status', 'active')
    with get_db() as db:
        db.execute("UPDATE users SET status=? WHERE id=?", (new_status, uid))
        db.commit()
    return jsonify({'success': True})


# ===================== 设备管理 API =====================

@app.route('/api/devices')
@login_required
def api_devices_list():
    """设备台账列表，支持按站点/类型/状态筛选"""
    site_id = request.args.get('site_id', '').strip()
    device_type = request.args.get('type', '').strip()
    status = request.args.get('status', '').strip()
    search = request.args.get('search', '').strip()
    with get_db() as db:
        sql = """SELECT d.id, d.device_code, d.device_name, d.device_type, d.status,
                        d.battery, d.voltage, d.last_data_time,
                        s.name as site_name, s.code as site_code, s.id as site_id
                 FROM device_shadows d LEFT JOIN sites s ON d.site_id=s.id WHERE 1=1"""
        params = []
        if site_id:
            sql += " AND d.site_id=?"
            params.append(site_id)
        if device_type:
            sql += " AND d.device_type=?"
            params.append(device_type)
        if status:
            sql += " AND d.status=?"
            params.append(status)
        if search:
            sql += " AND (d.device_name LIKE ? OR d.device_code LIKE ? OR s.name LIKE ?)"
            like = f"%{search}%"
            params.extend([like, like, like])
        sql += " ORDER BY d.status DESC, d.site_id"
        rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/devices/<int:device_id>')
@login_required
def api_device_detail(device_id):
    """设备详情 + 维护记录"""
    with get_db() as db:
        dev = db.execute("""SELECT d.*, s.name as site_name, s.code as site_code,
                                   s.district, s.manager
                            FROM device_shadows d LEFT JOIN sites s ON d.site_id=s.id
                            WHERE d.id=?""", (device_id,)).fetchone()
        if not dev:
            return jsonify({'error': '设备不存在'}), 404
        logs = db.execute("""SELECT * FROM inventory_logs
                             WHERE ref_type='maintenance' AND ref_id=?
                             ORDER BY created_at DESC LIMIT 20""", (device_id,)).fetchall()
    return jsonify({'device': dict(dev), 'logs': [dict(l) for l in logs]})


# --- 备件库存 ---

@app.route('/api/parts/inventory')
@login_required
def api_parts_inventory():
    """备件库存列表"""
    category = request.args.get('category', '').strip()
    low = request.args.get('low', '').strip()
    search = request.args.get('search', '').strip()
    with get_db() as db:
        sql = """SELECT p.*, s.name as site_name
                 FROM spare_parts_inventory p LEFT JOIN sites s ON p.site_id=s.id WHERE 1=1"""
        params = []
        if category:
            sql += " AND p.category=?"
            params.append(category)
        if low == '1':
            sql += " AND p.quantity <= p.min_quantity"
        if search:
            sql += " AND (p.part_name LIKE ? OR p.part_code LIKE ?)"
            like = f"%{search}%"
            params.extend([like, like])
        sql += " ORDER BY p.quantity ASC"
        rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/parts/inventory', methods=['POST'])
@login_required
def api_parts_inventory_add():
    """新增备件或入库"""
    data = request.get_json() or {}
    part_code = data.get('part_code', '').strip()
    part_name = data.get('part_name', '').strip()
    category = data.get('category', '其他').strip()
    unit = data.get('unit', '个').strip()
    quantity = int(data.get('quantity', 1))
    min_quantity = int(data.get('min_quantity', 5))
    site_id = data.get('site_id')
    remark = data.get('remark', '').strip()
    if not part_name:
        return jsonify({'error': '备件名称不能为空'}), 400
    if not part_code:
        import uuid
        part_code = f"BJ-{uuid.uuid4().hex[:6].upper()}"
    with get_db() as db:
        cur = db.execute("""INSERT INTO spare_parts_inventory
            (part_code,part_name,category,unit,quantity,min_quantity,site_id,remark)
            VALUES (?,?,?,?,?,?,?,?)""",
            (part_code, part_name, category, unit, quantity, min_quantity, site_id, remark))
        pid = cur.lastrowid
        db.execute("""INSERT INTO inventory_logs (part_id,type,quantity,ref_type,operator,remark)
            VALUES (?,'in',?,'purchase',?,?)""",
            (pid, quantity, g.current_user['username'] or 'admin', f'入库: {part_name}'))
        db.commit()
    return jsonify({'success': True, 'id': pid, 'part_code': part_code})


@app.route('/api/parts/inventory/<int:pid>', methods=['PUT'])
@login_required
def api_parts_inventory_update(pid):
    """更新备件信息或手动出库"""
    data = request.get_json() or {}
    with get_db() as db:
        part = db.execute("SELECT * FROM spare_parts_inventory WHERE id=?", (pid,)).fetchone()
        if not part:
            return jsonify({'error': '备件不存在'}), 404
        # 出库操作
        if 'out_qty' in data:
            qty = int(data['out_qty'])
            if qty <= 0:
                return jsonify({'error': '出库数量需大于0'}), 400
            if part['quantity'] - qty < 0:
                return jsonify({'error': '库存不足'}), 400
            db.execute("UPDATE spare_parts_inventory SET quantity=quantity-?, updated_at=datetime('now','localtime') WHERE id=?", (qty, pid))
            db.execute("""INSERT INTO inventory_logs (part_id,type,quantity,ref_type,operator,remark)
                VALUES (?,'out',?,'adjust',?,?)""",
                (pid, qty, g.current_user['username'] or 'admin', '手动出库'))
        # 入库操作
        if 'in_qty' in data:
            qty = int(data['in_qty'])
            if qty <= 0:
                return jsonify({'error': '入库数量需大于0'}), 400
            db.execute("UPDATE spare_parts_inventory SET quantity=quantity+?, updated_at=datetime('now','localtime') WHERE id=?", (qty, pid))
            db.execute("""INSERT INTO inventory_logs (part_id,type,quantity,ref_type,operator,remark)
                VALUES (?,'in',?,'purchase',?,?)""",
                (pid, qty, g.current_user['username'] or 'admin', '手动入库'))
        # 更新基本信息
        for field in ['part_name', 'category', 'unit', 'min_quantity', 'remark']:
            if field in data:
                db.execute(f"UPDATE spare_parts_inventory SET {field}=? WHERE id=?", (data[field], pid))
        db.commit()
    return jsonify({'success': True})


@app.route('/api/parts/inventory/<int:pid>/logs')
@login_required
def api_parts_inventory_logs(pid):
    """备件库存变更流水"""
    with get_db() as db:
        rows = db.execute("""SELECT * FROM inventory_logs WHERE part_id=? ORDER BY created_at DESC LIMIT 50""", (pid,)).fetchall()
    return jsonify([dict(r) for r in rows])


# --- 备件申请 ---

@app.route('/api/parts/requests', methods=['GET'])
@login_required
def api_parts_requests_list():
    """备件申请列表（Web端：全部；移动端按申请人过滤）"""
    status_f = request.args.get('status', '').strip()
    applicant = request.args.get('applicant', '').strip()
    with get_db() as db:
        sql = """SELECT r.*, s.name as site_name
                 FROM spare_part_requests r LEFT JOIN sites s ON r.site_id=s.id WHERE 1=1"""
        params = []
        if status_f:
            sql += " AND r.status=?"
            params.append(status_f)
        if applicant:
            sql += " AND r.applicant=?"
            params.append(applicant)
        sql += " ORDER BY r.created_at DESC"
        rows = db.execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/parts/requests', methods=['POST'])
@login_required
def api_parts_requests_create():
    """创建备件申请（移动端调用）"""
    data = request.get_json() or {}
    site_id = data.get('site_id')
    part_name = data.get('part_name', '').strip()
    quantity = int(data.get('quantity', 1))
    reason = data.get('reason', '').strip()
    if not site_id or not part_name:
        return jsonify({'error': '站点和备件名称不能为空'}), 400
    applicant = g.current_user['username'] or 'unknown'
    # 生成申请编号：BJ+年月日+序号
    from datetime import datetime
    today = datetime.now().strftime('%Y%m%d')
    with get_db() as db:
        count = db.execute("SELECT COUNT(*) as c FROM spare_part_requests WHERE request_no LIKE ?", (f"BJ-{today}%",)).fetchone()['c']
        request_no = f"BJ-{today}-{count+1:03d}"
        db.execute("""INSERT INTO spare_part_requests
            (request_no,site_id,applicant,part_name,quantity,reason)
            VALUES (?,?,?,?,?,?)""",
            (request_no, site_id, applicant, part_name, quantity, reason))
        db.commit()
    return jsonify({'success': True, 'request_no': request_no})


@app.route('/api/parts/requests/mine')
@login_required
def api_parts_requests_mine():
    """我的备件申请记录（移动端）"""
    applicant = g.current_user['username'] or 'unknown'
    with get_db() as db:
        rows = db.execute("""SELECT r.*, s.name as site_name
            FROM spare_part_requests r LEFT JOIN sites s ON r.site_id=s.id
            WHERE r.applicant=? ORDER BY r.created_at DESC""", (applicant,)).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/parts/requests/<int:rid>/approve', methods=['PUT'])
@login_required
def api_parts_request_approve(rid):
    """审批通过：更新状态 + 扣减库存"""
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '仅管理员可审批'}), 403
    comment = (request.get_json() or {}).get('comment', '审批通过')
    with get_db() as db:
        req = db.execute("SELECT * FROM spare_part_requests WHERE id=?", (rid,)).fetchone()
        if not req:
            return jsonify({'error': '申请不存在'}), 404
        if req['status'] != 'pending':
            return jsonify({'error': '该申请已处理'}), 400
        # 更新申请状态
        db.execute("""UPDATE spare_part_requests SET status='approved', approver=?,
            approval_comment=?, updated_at=datetime('now','localtime') WHERE id=?""",
            (g.current_user['username'] or 'admin', comment, rid))
        # 尝试扣减库存：查找匹配的备件（按名称模糊匹配）
        inv = db.execute("""SELECT * FROM spare_parts_inventory
            WHERE part_name LIKE ? ORDER BY quantity DESC LIMIT 1""",
            (f"%{req['part_name']}%",)).fetchone()
        if inv:
            new_qty = max(0, inv['quantity'] - req['quantity'])
            db.execute("UPDATE spare_parts_inventory SET quantity=?, updated_at=datetime('now','localtime') WHERE id=?", (new_qty, inv['id']))
            db.execute("""INSERT INTO inventory_logs (part_id,type,quantity,ref_type,ref_id,operator,remark)
                VALUES (?,'out',?,'request',?,?,?)""",
                (inv['id'], req['quantity'], rid, g.current_user['username'] or 'admin',
                 f"备件申请 #{req['request_no']}"))
        db.commit()
    return jsonify({'success': True, 'message': '已批准，库存已扣减'})


@app.route('/api/parts/requests/<int:rid>/reject', methods=['PUT'])
@login_required
def api_parts_request_reject(rid):
    """驳回申请"""
    if g.current_user['role'] != 'admin':
        return jsonify({'error': '仅管理员可审批'}), 403
    comment = (request.get_json() or {}).get('comment', '驳回')
    with get_db() as db:
        req = db.execute("SELECT * FROM spare_part_requests WHERE id=?", (rid,)).fetchone()
        if not req:
            return jsonify({'error': '申请不存在'}), 404
        if req['status'] != 'pending':
            return jsonify({'error': '该申请已处理'}), 400
        db.execute("""UPDATE spare_part_requests SET status='rejected', approver=?,
            approval_comment=?, updated_at=datetime('now','localtime') WHERE id=?""",
            (g.current_user['username'] or 'admin', comment, rid))
        db.commit()
    return jsonify({'success': True, 'message': '已驳回'})


# ===================== 站点过滤辅助函数 =====================

def _filter_by_user(where_clause='', table_prefix=''):
    """为API查询注入站点过滤条件。返回 (where_extra, params)
    管理员不限制，操作员限制为分配的站点。
    在路由函数中使用：在原始WHERE后加上此函数的返回。
    """
    site_ids = getattr(g, 'user_site_ids', None)
    if site_ids is None:
        return '', []
    prefix = table_prefix + '.' if table_prefix else ''
    site_condition = f"{prefix}site_id IN ({','.join('?' * len(site_ids))})" if site_ids else '1=0'
    extra = f" AND {site_condition}" if where_clause else f" WHERE {site_condition}"
    return extra, site_ids


def _filter_site_ids():
    """返回当前用户可见的site_id列表（管理员返回空=全部）"""
    site_ids = getattr(g, 'user_site_ids', None)
    if site_ids is None:
        return None  # None = 不过滤
    return site_ids


# ===================== 前端静态文件服务 =====================
FRONTEND_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'frontend')

@app.route('/')
def index_html():
    return send_from_directory(FRONTEND_DIR, 'dashboard.html')

# 移动端（放在catch-all前面）
@app.route('/mobile')
def mobile_page():
    return send_from_directory(FRONTEND_DIR, 'mobile.html')

@app.route('/<path:filename>')
def serve_frontend(filename):
    return send_from_directory(FRONTEND_DIR, filename)

# ===================== Startup =====================

if __name__ == '__main__':
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    init_db()
    seed_data()
    seed_inspections()
    seed_alerts()
    seed_maintenance()
    seed_maintenance_templates()
    seed_users()
    backfill_history(72)
    # 生成初始数据（让趋势跟踪生效）
    for _ in range(6):
        try:
            if os.environ.get('SKIP_SEED') != '1':
                generate_sensor_data()
        except Exception as e:
            print(f'[Seed] 初始数据生成跳过: {e}')
    # 预置少量固定离线站点（在初始数据生成之后执行，确保不被覆盖）
    try:
        with get_db() as db:
            for sid in [5, 108, 193]:
                db.execute("UPDATE device_shadows SET status='offline', last_data_time=NULL WHERE site_id=?", (sid,))
                site = db.execute("SELECT name FROM sites WHERE id=?", (sid,)).fetchone()
                if site:
                    print(f"[Seed] 预设离线站点: {site['name']} (id={sid})")
                    # 生成离线告警
                    devs = db.execute("SELECT device_name, device_code FROM device_shadows WHERE site_id=?", (sid,)).fetchall()
                    for dev in devs:
                        create_alert_internal(db, sid, 'device_status', 0, 'yellow',
                            f"设备离线: {dev['device_name']} ({dev['device_code']}) · {site['name']}")
            db.commit()
    except Exception as e:
        print(f"[Seed] 预设离线站点跳过: {e}")
    # 清理过量的已办结告警（只保留最近1天的）
    try:
        with get_db() as db:
            total = db.execute("SELECT COUNT(*) as c FROM alerts WHERE status='resolved'").fetchone()['c']
            if total > 200:
                db.execute("DELETE FROM alerts WHERE status='resolved' AND created_at < datetime('now','-1 day')")
                db.commit()
                print(f"[Cleanup] 清理过量已办结告警: 删除前{total}条，保留最近1天记录")
    except Exception as e:
        print(f"[Cleanup] 告警清理跳过: {e}")
    # 每30秒自动生成数据
    scheduler.add_job(generate_sensor_data, 'interval', seconds=60, id='simulator')
    # 每30分钟更新天气
    scheduler.add_job(fetch_real_weather, 'interval', minutes=30, id='weather_updater')
    print("[Server] 水利运维智慧运营平台 启动成功!")
    print("[Server] API: http://localhost:5000/api/health")
    app.run(host='0.0.0.0', port=5000, debug=False, threaded=True)
